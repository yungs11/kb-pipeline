"""csv 바이트 → xlsx 바이트. 엑셀 레인이 csv 를 받기 위한 얇은 어댑터.

**헤더 행에 서식(볼드 + 채우기)을 반드시 준다.** `excel_parser_rag` 의 헤더 감지는
`strong` 판정에 `eff_style >= _STYLE_GATE_MIN` 을 요구한다
(`excel_parser_rag/detection/header_detector.py:318-323`). 서식 없는 맨 셀로 합성하면
감지가 **구조적으로 실패**하고 `key = headers.get(c) or get_column_letter(c)`
(`parsers/flat_table.py:175`) 폴백이 걸려 청크가 `사번: 1001` 이 아니라 `A: 1001` 로
퇴화한다. 헤더 행 자체도 데이터행으로 오인돼 청크가 하나 늘어난다(실측).

**openpyxl 이 거부하는 입력 둘을 미리 막는다.** 시트명의 ``[ ] : * ? / \\`` 는
title 대입에서 ValueError, 셀의 제어문자는 값 대입에서 IllegalCharacterError 를 낸다.
둘 다 excel/__init__.py 의 `except Exception` 을 타고 ParserError 로 승격돼 **문서 전체가
parse_failed** 가 된다 — `현황[최종].csv` 같은 파일명은 지금 text 레인으로 잘 통과하므로
막지 않으면 순수 회귀다.

**모든 문자열 셀의 data_type 을 's' 로 고정한다.** 안 하면 ``=`` 로 시작하는 셀이 수식
셀(``data_type='f'``)이 되는데, 엑셀 레인은 ``data_only=True`` 로 읽으므로 캐시된 계산값이
없는 그 셀은 **None** 이 되어 청크에서 통째로 사라진다. 게이트는 ok=True 로 통과시켜
**조용한 데이터 손실이 적재까지 간다**(실측: ``[['사번','수식'],['1001',None],…]``).
csv 는 애초에 전부 텍스트라 고정에 부작용이 없고, CSV 인젝션 방어도 겸한다.

**첫 행을 헤더로 간주한다**(csv 포맷 관례). 헤더 없는 파일이면 첫 데이터 행이 컬럼명이
된다 — 헤더 유무 감지는 범위 밖(`deferred.md` D46).
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from parse_service.parsers import ParserError
from parse_service.tools.textdecode import decode_text

#: openpyxl 시트명 상한. 넘기면 저장이 아니라 title 대입에서 바로 터진다.
_SHEET_TITLE_MAX = 31
#: openpyxl 이 시트명에서 거부하는 문자.
_SHEET_BAD_CHARS = re.compile(r"[\\/*?:\[\]]")


def _sheet_title(filename: str) -> str:
    stem = _SHEET_BAD_CHARS.sub("_", Path(filename).stem)
    return stem[:_SHEET_TITLE_MAX] or "Sheet1"


def csv_bytes_to_xlsx(file_bytes: bytes, filename: str) -> bytes:
    text = decode_text(file_bytes, filename)
    # newline="" 이어야 인용 셀 안의 개행이 보존된다(csv 모듈 문서의 요구사항).
    rows = [r for r in csv.reader(io.StringIO(text, newline=""))
            if any((c or "").strip() for c in r)]
    if not rows:
        raise ParserError(f"empty csv: {filename}")

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(filename)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx,
                           value=ILLEGAL_CHARACTERS_RE.sub("", value or ""))
            # 수식 승격 차단 — 위 docstring 참조. append() 대신 cell() 을 쓰는 이유가
            # 이것이다(append 로는 대입 직후 data_type 을 잡을 지점이 없다).
            cell.data_type = "s"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
