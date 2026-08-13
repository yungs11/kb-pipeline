"""Workbook 로더 (SoT §4 step 1~2, §5.1, §35).

pipeline.py 가 고정한 시그니처:

- load_workbook_for_parsing(path, config) -> (data_wb, formula_wb|None)
- should_skip_sheet(ws, config) -> bool
"""

from __future__ import annotations

import tempfile
import warnings
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..config import ParserConfig
from .xls_converter import convert_xls_to_xlsx


def load_workbook_for_parsing(
    path: str | Path,
    config: ParserConfig,
) -> Tuple[Workbook, Optional[Workbook]]:
    """엑셀 파일을 (캐시값 워크북, 수식 워크북|None) 으로 로드한다.

    - data_only=True: 수식 셀은 캐시된 결과값으로 읽는다 (formula_mode=cached_value).
    - config.formula_mode == "formula_text" 면 data_only=False 워크북도 함께 로드해
      수식 텍스트를 CellNode.formula 로 제공할 수 있게 한다.
    - .xls 입력은 xls_converter 로 .xlsx 변환 후 처리한다.
    - openpyxl 의 UserWarning(미지원 확장, 기본 스타일 누락 등)은 억제한다.
    """
    src = Path(path)
    if not src.is_file():
        raise FileNotFoundError(f"입력 파일이 존재하지 않습니다: {src}")
    if src.suffix.lower() == ".xls":
        # 레인 입구(`parsers/excel/__init__.py`)가 CFB 를 이미 갈아끼우므로 실전에서는
        # 여기 `.xls` 가 도달하지 않는다. 그래도 **output_dir 을 넘긴다** — 같은 함수의
        # 호출 규약이 호출처마다 다르면 다음 사람이 또 밟는다(미지정 시 변환기가 만드는
        # 임시 디렉터리는 지우는 주체가 없어 사본이 영구 누적된다).
        src = convert_xls_to_xlsx(src, output_dir=tempfile.mkdtemp(prefix="xls2xlsx_"))

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        data_wb = openpyxl.load_workbook(src, data_only=True)
        formula_wb: Optional[Workbook] = None
        if config.formula_mode == "formula_text":
            formula_wb = openpyxl.load_workbook(src, data_only=False)
    return data_wb, formula_wb


def should_skip_sheet(ws: Worksheet, config: ParserConfig) -> bool:
    """파싱에서 제외할 시트인지 판단한다.

    - sheet_overrides 에서 skip=True 로 지정된 시트
    - 숨김 시트 (config.parse_hidden_sheets=False 인 경우)
    """
    override = config.sheet_overrides.get(ws.title)
    if override is not None and override.skip:
        return True
    sheet_state = getattr(ws, "sheet_state", "visible")
    if sheet_state != "visible" and not config.parse_hidden_sheets:
        return True
    return False
