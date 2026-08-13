"""게이트 검증 요약 — 추출 실패/불명확 표 구조.

설계: docs/superpowers/specs/2026-06-29-excel-gate-postparse-design.md
백엔드(openpyxl/kordoc) 무관하게 동작: 원시 셀(openpyxl)로 구조/참조,
실제 파싱 chunks 로 헤더누수를 판정한다.
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from ..config import ParserConfig
from ..pipeline import build_canvases, detect_and_classify
from ..parsers.flat_table import cell_text, body_rows_of
from ..chunking.sibling_rule import _is_meta_sheet  # 크로스모듈 재사용(순환 없음 — 스텝 0 확인)

# unclear_header 후보 리전 타입 — 헤더 없이 표류하는 표(전결 Index/자산목록 등과 동형).
_UNCLEAR_DRIFT_TYPES = {"flat_table", "unknown_table", "code_mapping_table", "key_value_block"}

# ★ IGNORECASE 필수(2026-08-13 실측) — **LibreOffice 는 캐시된 오류값을 소문자로 쓴다.**
# 레거시 .xls 를 soffice 로 .xlsx 변환해 들여오면 `#REF!` 가 `#ref!` 가 되고, 수식 쪽도
# `=#REF!` → `="#ref!"`(문자열 리터럴)로 바뀐다. 대문자 전용 패턴이면 값·수식 양쪽 스캔이
# 전부 빗나가 **참조오류 게이트가 .xls 문서에서 통째로 침묵**한다(= 불량 문서가 통과).
# 소문자 `#ref!` 가 정상 본문에 나올 일은 사실상 없어 오탐 위험은 무시할 수준이다.
ERROR_RE = re.compile(r"#(REF|VALUE|DIV/0|N/A|NAME\?|NULL|NUM)!?", re.IGNORECASE)
# 약어 의미 정규화 — 공백·각주 기호 제거('기술관리실*' vs '기술관리실' 동일 취급, 비상충).
# 주의: internal whitespace 도 제거됨('경영 기획팀'=='경영기획팀') — 미발화 쪽 오류라 정밀-안전.
_MEANING_NORM_RE = re.compile(r"[\s*※†‡]+")


def _conflicting_codes(chunks: List[Dict[str, Any]]) -> Dict[str, Dict[str, List[str]]]:
    """code_mapping 청크에서 sheet → {약어: 상충 원문 의미들(2종+만)}.

    병합 fields['매핑'](리스트) 우선, 부재 시 낱개 {약어,의미} fallback —
    merge_max_chars=0 롤백 상태에서도 게이트가 독립 동작(게이트-병합 디커플링).
    """
    seen: Dict[Tuple[str, str], Dict[str, str]] = {}
    for c in chunks:
        if c.get("chunk_type") != "code_mapping":
            continue
        f = c.get("fields") or {}
        pairs = f.get("매핑")
        if not isinstance(pairs, list):            # 낱개(롤백) fallback
            pairs = [{"약어": f.get("약어"), "의미": f.get("의미")}]
        for m in pairs:
            code = str(m.get("약어") or "").strip()
            meaning = str(m.get("의미") or "").strip()
            if not code or not meaning:
                continue
            norm = _MEANING_NORM_RE.sub("", meaning)
            seen.setdefault((c.get("sheet", ""), code), {}).setdefault(norm, meaning)
    out: Dict[str, Dict[str, List[str]]] = {}
    for (s, code), v in seen.items():
        if len(v) > 1:
            out.setdefault(s, {})[code] = sorted(v.values())
    return out


def _header_labels(region, canvas) -> Dict[int, str]:
    """region 헤더행의 {col: label}. header_rows 없으면 빈 dict."""
    out: Dict[int, str] = {}
    for hr in (region.header_rows or []):
        for col in range(region.min_col, region.max_col + 1):
            cell = canvas.cells.get((hr, col))
            # CellNode 필드명: display_value / normalized_value / logical_value (cell_node.py).
            # 병합·복원 셀까지 잡으려면 logical_value 우선.
            val = "" if cell is None else ("" if cell.is_empty else str(getattr(cell, "logical_value", "") or cell.display_value or cell.normalized_value or "").strip())
            if val and col not in out:
                out[col] = val
    return out


def _region_width_banner_rows(region, canvas) -> List[int]:
    """리전 사용 열 전체(min_col..max_col)를 덮는 병합 배너 행(텍스트 非빈, 폭≥2)의 행번호.

    게이트 로컬 헬퍼(★유일·권위 분리자, v2). hierarchy_table.is_full_width_banner 는 width<3
    가드가 있어 2열 리전(R&R A:B)에서 부적합하므로 무수정 존치하고 여기서 별도 판정한다.
    빈 병합 구분행(텍스트 빈칸)은 배너 아님(v2-b): cell_text 비어있으면 제외.
    """
    width = region.max_col - region.min_col + 1
    if width < 2:
        return []
    rows: List[int] = []
    for mr in canvas.merged_ranges:
        try:
            c0, r0, c1, r1 = range_boundaries(mr)
        except Exception:
            continue
        if not (region.min_row <= r0 <= region.max_row):
            continue
        if c0 <= region.min_col and c1 >= region.max_col and (c1 - c0 + 1) >= 2:
            if cell_text(canvas.get_cell(r0, c0)).strip():
                rows.append(r0)
    return rows


def _unmerged_singleton_title_rows(region, canvas) -> List[int]:
    """병합 없는 넓은 리전에서 왼쪽 단일 셀만 채운 제목 후보 행.

    단일값 행만으로 제목/데이터를 확정할 수는 없으므로 이 헬퍼는 후보만 반환한다.
    호출부가 ``초기 제목 + 헤더 이후 재등장 + 2회 이상``을 함께 요구해 오탐을 줄인다.
    리전과 겹치는 병합이 하나라도 있으면 문서 작성자가 표 경계를 구조화한 것으로 보고 제외한다.
    """
    if region.max_col - region.min_col + 1 < 3:
        return []

    for mr in canvas.merged_ranges:
        try:
            c0, r0, c1, r1 = range_boundaries(mr)
        except Exception:
            continue
        overlaps = not (
            r1 < region.min_row or r0 > region.max_row
            or c1 < region.min_col or c0 > region.max_col
        )
        if overlaps:
            return []

    rows: List[int] = []
    for r in range(region.min_row, region.max_row + 1):
        values = [
            (c, cell_text(canvas.get_cell(r, c)).strip())
            for c in range(region.min_col, region.max_col + 1)
        ]
        values = [(c, value) for c, value in values if value]
        if len(values) != 1 or values[0][0] != region.min_col:
            continue
        value = values[0][1]
        # 번호/날짜/기호만 있는 일반 데이터 파편은 제목 후보에서 제외.
        if len(value) > 80 or not re.search(r"[A-Za-z가-힣]", value):
            continue
        rows.append(r)
    return rows


def _empty_split_block_first_rows(region, canvas) -> List[int]:
    """리전을 빈 행으로 분할한 각 non-empty 블록의 '첫 행' 목록.

    블록마다 다른 첫 행이 각각 헤더로 승격되는 구조 신호(≥2 블록이면 데이터-헤더 오인 위험).
    """
    firsts: List[int] = []
    cur = False
    for r in range(region.min_row, region.max_row + 1):
        ne = any(cell_text(canvas.get_cell(r, c)).strip()
                 for c in range(region.min_col, region.max_col + 1))
        if ne and not cur:
            firsts.append(r)
            cur = True
        elif not ne:
            cur = False
    return firsts


def _block_first_cell(region, canvas, r) -> str:
    """블록 첫 행 r 의 대표 셀(왼쪽 첫 non-empty) A1 좌표."""
    for c in range(region.min_col, region.max_col + 1):
        if cell_text(canvas.get_cell(r, c)).strip():
            return f"{get_column_letter(c)}{r}"
    return f"{get_column_letter(region.min_col)}{r}"


def compute_gate_summary(input_path, chunks: List[Dict[str, Any]]) -> Dict[str, Any]:
    path = Path(input_path)
    cfg = ParserConfig()
    canvases = build_canvases(path, cfg)
    region_pairs = detect_and_classify(canvases, cfg)

    # region 을 시트별로 묶기
    by_sheet: Dict[str, list] = defaultdict(list)
    for region, canvas in region_pairs:
        by_sheet[canvas.sheet_name].append((region, canvas))

    # 원시 워크북(참조오류 스캔용).
    #  - data_only=True : 캐시된 계산 결과의 에러 문자열(#REF! 등)
    #  - data_only=False: 수식 문자열 자체의 깨진 참조(=SUM(#REF!) 등) — 캐시가 없을 때도 잡는다.
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        wb_formula = openpyxl.load_workbook(path)  # 기본=수식 보존
    except Exception:
        wb_formula = None

    conflicts = _conflicting_codes(chunks)  # sheet → {약어: 상충 의미들}. 루프 밖 1회.

    sheets_out: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        findings: List[Dict[str, Any]] = []

        # 1) ref_error — 값(캐시) + 수식 문자열 양쪽 스캔
        ref_set: set = set()
        ws_f = wb_formula[ws.title] if (wb_formula is not None and ws.title in wb_formula.sheetnames) else None
        for src_ws in (ws, ws_f):
            if src_ws is None:
                continue
            for row in src_ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and ERROR_RE.search(cell.value):
                        ref_set.add((cell.row, cell.column))
        ref_cells = [f"{get_column_letter(c)}{r}" for r, c in sorted(ref_set)]
        if ref_cells:
            # 플래그 좌표마다 무조건 캐시값(#REF!)+수식(=J9)을 조회.
            #   - H3=`=J9` 처럼 에러토큰 없는 수식셀은 캐시 #REF! 로만 플래그되므로
            #     ERROR_RE 매치 분기 안에서 수식을 잡으면 도달 불가 → 좌표 무조건 조회.
            #   - MergedCell 방어: getattr 로 value 안전 접근.
            parts: List[str] = []
            for r, c in sorted(ref_set):
                coord = f"{get_column_letter(c)}{r}"
                value = getattr(ws.cell(r, c), "value", None)
                formula = getattr(ws_f.cell(r, c), "value", None) if ws_f is not None else None
                fml = str(formula)[:60] if formula is not None else None
                if value is None and fml:
                    parts.append(f"{coord}=({fml})")
                elif fml:
                    parts.append(f"{coord}={value} ({fml})")
                else:
                    parts.append(f"{coord}={value}")
                if len(parts) >= 5:
                    break
            findings.append({"code": "ref_error", "cells": ref_cells[:20],
                             "detail": f"참조 오류가 값에 포함됨: {', '.join(parts)}"})

        # 2) empty_header — region 헤더 기반.
        # side_by_side 규칙은 정상 병렬 표까지 차단하므로 사용자 결정으로 비활성화했다.
        for region, canvas in by_sheet.get(ws.title, []):
            labels = _header_labels(region, canvas)
            if not labels:
                continue
            # empty_header: 사용 열에 헤더 라벨이 비어있는 칸.
            # ⚠️ 의도적으로 거의 비활성(보수적): region 이 잡혔다는 건 보통 라벨 ≥2 이므로
            #   아래 `len(labels) < 2` 게이트로 실질적 미발화. trailing blank column 오탐을
            #   피하려는 스캐폴딩이며, 추후 명확한 트리거 정의가 생기면 완화한다.
            # 임계치: 전체 region 열 중 빈 헤더 비율 50% 초과 AND 라벨<2 일 때만.
            total_cols = region.max_col - region.min_col + 1
            empty_cols = [get_column_letter(col) + str(region.header_rows[0])
                          for col in range(region.min_col, region.max_col + 1) if col not in labels]
            empty_ratio = len(empty_cols) / total_cols if total_cols > 0 else 0
            if empty_cols and empty_ratio > 0.5 and len(labels) < 2:
                findings.append({"code": "empty_header", "cells": empty_cols[:20],
                                 "detail": f"헤더 컬럼명이 비어있음: {', '.join(empty_cols[:5])}"})

        # ambiguous_hierarchy는 서로 다른 열의 정상적인 번호 재시작까지 차단하므로
        # 사용자 결정으로 비활성화했다. 실제 같은-열 스택 충돌 사례가 확보되기 전에는
        # 번호 패턴만으로 계층 모호성을 추정하지 않는다.

        # 3c) conflicting_code_mapping — 같은 약어가 상충 의미로 다중 정의(개정 전/후 대조 블록 등)
        sheet_conf = conflicts.get(ws.title) or {}
        if sheet_conf:
            codes = sorted(sheet_conf)
            sample = "; ".join(f"'{c}': {', '.join(sheet_conf[c][:3])}" for c in codes[:3])
            findings.append({
                "code": "conflicting_code_mapping",
                "cells": codes[:20],   # 셀좌표 대신 약어 — 사람이 바로 찾음(doc_guard list[str] join)
                "detail": f"약어 {len(codes)}건이 서로 다른 의미로 중복 정의됨 (예: {sample})",
            })

        # 3d) unclear_header — 헤더 행 자체가 없는 표(첫 데이터 행이 헤더로 오인).
        #     empty_header("헤더행은 있고 라벨 칸이 빔")와 구분되는 신규 신호. region 기반(backend 무관).
        #     ★별도 리전 루프 — empty_header 루프는 labels 없으면
        #     early-continue 라 그 안에 넣으면 영원히 미발화(v2-b). 신호 전부 AND:
        #       (1) 표류 리전(flat/unknown/code_mapping/key_value) ∧ header_rows 미검출 ∧ body≥4
        #       (2) 비메타 시트  (3) 빈행분할 블록≥2  (4) 리전폭 병합배너 존재(유일·권위 분리자)
        #     실측 분리: R&R(배너 有)만 발화 / AI Agent 참고표·Index 약어표(배너 無) 미발화.
        if not _is_meta_sheet(ws.title):
            for region, canvas in by_sheet.get(ws.title, []):
                if region.region_type not in _UNCLEAR_DRIFT_TYPES or region.header_rows:
                    continue
                if len(body_rows_of(region, canvas)) < 4:
                    continue
                block_firsts = _empty_split_block_first_rows(region, canvas)
                if len(block_firsts) < 2:
                    continue
                if not _region_width_banner_rows(region, canvas):
                    continue
                cells = [_block_first_cell(region, canvas, r) for r in block_firsts]
                findings.append({
                    "code": "unclear_header",
                    "cells": cells[:20],
                    "detail": "표에 열 제목(헤더) 행이 없어 첫 데이터 행이 헤더로 오인될 수 있음",
                })
                break  # 시트당 1건(관례)

        # 3e) unmerged_table_banners — 넓은 표 안에 병합 없는 왼쪽 단일 셀 제목 행으로
        #     여러 표를 세로 적층한 구조. 스타일 없는 백엔드에서는 제목/데이터를 구분할
        #     권위 신호가 없으므로 파서가 임의 분할하지 않고 문서 수정을 안내한다.
        #     정밀 우선 신호 전부 AND:
        #       (1) 비메타, 폭≥3, 헤더 라벨≥2, body≥4
        #       (2) 리전 겹침 병합 0
        #       (3) 단일 셀 제목 후보≥2
        #       (4) 후보가 최초 헤더 이전/동일에 1개 이상 + 최종 헤더 이후에 1개 이상
        if not _is_meta_sheet(ws.title):
            for region, canvas in by_sheet.get(ws.title, []):
                if not region.header_rows or len(body_rows_of(region, canvas)) < 4:
                    continue
                if len(_header_labels(region, canvas)) < 2:
                    continue
                title_rows = _unmerged_singleton_title_rows(region, canvas)
                if len(title_rows) < 2:
                    continue
                if not any(r <= min(region.header_rows) for r in title_rows):
                    continue
                if not any(r > max(region.header_rows) for r in title_rows):
                    continue
                cells = [f"{get_column_letter(region.min_col)}{r}" for r in title_rows]
                findings.append({
                    "code": "unmerged_table_banners",
                    "cells": cells[:20],
                    "detail": (
                        "병합되지 않은 단일 셀 제목 행 때문에 여러 표의 표 경계가 불명확함; "
                        "각 표 제목을 표 너비로 병합하고 각 표에 헤더 행을 추가해야 함"
                    ),
                })
                break  # 시트당 1건(관례)

        # 4) header_leak — chunk 의 field[k]==k (헤더행이 데이터로 추출됨)
        for c in chunks:
            if c.get("sheet") != ws.title:
                continue
            fields = c.get("fields") or {}
            if not isinstance(fields, dict) or len(fields) < 2:
                continue
            same = sum(1 for k, v in fields.items()
                       if isinstance(v, str) and v.strip() == str(k).strip() and v.strip() != "")
            if same >= max(2, (len(fields) + 1) // 2):
                src = c.get("source") or {}
                row = src.get("start_row")
                loc = [f"row{row}"] if row else []
                leaked = [str(k) for k, v in fields.items()
                          if isinstance(v, str) and v.strip() == str(k).strip() and v.strip()][:5]
                pairs = ", ".join(f"{k}={k}" for k in leaked)
                # row{n}(청크 데이터행)은 실제 헤더행과 다를 수 있어 모호 → 헤더=값 쌍으로 자기설명.
                findings.append({"code": "header_leak", "cells": loc,
                                 "detail": f"헤더행이 데이터로 추출됨(이 행이 헤더 라벨과 동일): {pairs}"})
                break  # 시트당 1건이면 충분

        sheets_out.append({"sheet": ws.title, "ok": not findings, "findings": findings})

    wb.close()
    if wb_formula is not None:
        wb_formula.close()
    return {"ok": all(s["ok"] for s in sheets_out), "sheets": sheets_out}
