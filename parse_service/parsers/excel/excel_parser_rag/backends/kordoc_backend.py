"""kordoc(.md) 기반 백엔드 (kordoc 통합설계).

흐름: xlsx → (kordoc .md 확보: 지정/디렉토리/자동생성) → HTML grid 복원 → 시트제목·배너
분할 → 섹션별 헤더 → 행 청크(compact matrix + table_row) → k2o 좌표 → SoT 호환 chunk dict.

md 확보 우선순위:
  1. config.kordoc_md_path (단일 파일)
  2. config.kordoc_md_dir/<stem>.md
  3. config.kordoc_bin 으로 자동 생성 (Node 필요) → config.kordoc_md_out (기본 임시)
없으면 명확히 에러(자동 openpyxl fallback 금지 — silent 품질저하 방지).
"""
from __future__ import annotations

import difflib
import os
import re
import shlex
import subprocess
import tempfile
from collections import Counter
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from ..config import ParserConfig
from ..textutil import (
    PARSER_VERSION,
    _is_pct_format,
    _pct_decimals,
    build_content_frame,
    infer_numbering_level,
    mark_strikethrough,
    range_a1,
)
from .base import BackendError

# ─── 헤더 스코어러 상수 ──────────────────────────────
_MAX_HEADER_SCAN = 8
_NUMERIC_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?%?$")
_FORMULA_ERR_RE = re.compile(r"#(REF|VALUE|DIV/0|N/A|NAME\?|NULL|NUM)!?")
# 순수 장식 세그먼트(구분선 '----', '====', '─' 등) — 제목/섹션/경로 세그로 잘못 승격되면
# kv 키·경로를 오염시키므로 정화한다. 전체-일치(^…$)라 통짜 장식 세그만 제거(혼합 텍스트 보존).
_DECOR_RE = re.compile(r"^[\s\-=─—–ㅡ_·・.*#~]+$")
# 부분병합 소제목(single_left) 승격 시 번호매김 없는 라벨의 최대 길이(짧은 라벨만 소제목).
_SUBTITLE_MAX_LEN = 40
# header-less 청크 의미 키 판정용: 주석 접두(마커 ● 는 is_marker 로 먼저 배제) + <별표N> 패턴.
_HL_NOTE_PREFIXES = ("※", "*", "☞", "▶", "◈", "■", "●", "◆")
_BYULPYO_RE = re.compile(r"^[<(\[]?\s*별표")

# 마커 열 게이트 임계: 열 내 마커비율(마커셀/nonblank 본문셀, 섹션 스코프)이 이 값 이상이면
# '마커 전용 열'로 보고 matrix_fact 지표로 인정한다. 코퍼스 11파일 실측 쌍봉 분포에서 도출:
#   전체자산 접근제어 col49 = 3.8%(6/157, 유일한 값 열) vs 전결표류 마커 열 30개 = 20~100%.
#   중간(4~20%) 공백 구간이 넓어 0.15 임계는 양쪽에 여유가 있다.
_MARKER_COL_MIN = 0.15

# ─── 마커 ────────────────────────────────────────────
MARKER_NORM = {
    "○": "applicable", "◯": "applicable", "●": "applicable_primary", "◎": "applicable_special",
    "△": "conditional", "▲": "conditional", "×": "not_applicable", "✕": "not_applicable",
    "✓": "checked", "✔": "checked", "√": "checked",
}
MARK_BUCKET = {
    "○": "해당", "◯": "해당", "●": "해당", "◎": "해당", "△": "조건부", "▲": "조건부",
    "×": "비해당", "✕": "비해당", "✓": "체크", "✔": "체크", "√": "체크",
}


def is_marker(t: Any) -> bool:
    return str(t or "").strip() in MARKER_NORM


def clean_title(t: Any) -> str:
    return re.sub(r"^[▶▷●■◆\s]+", "", str(t or "")).strip()


def _desecorate_title(title: str) -> str:
    """' / ' 조인 제목에서 순수장식 서브세그(구분선)만 드랍(보조 정화 — 조인 후 잔여 방어).

    Part A 의 1차 방어는 _segment 의 순수장식 행 skip 이고, 이 함수는 혹시 조인된 장식 꼬리
    (' / --------------------')가 남을 경우의 2차 방어다. 전체-일치 세그만 제거해 혼합 텍스트 보존.
    """
    if not title:
        return title
    kept = [s.strip() for s in title.split(" / ")
            if s.strip() and not _DECOR_RE.match(s.strip())]
    return " / ".join(kept)


def clean_val(t: Any) -> str:
    return re.sub(r"\s*\n\s*", " ", str(t)).strip()


def norm(s: Any) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


# ─── 날짜 시리얼 보정 (openpyxl 직접 순회, MergedCell 미접근) ─────────
def _fmt_dt(v):
    if isinstance(v, datetime) and (v.hour or v.minute or v.second):
        return v.strftime("%Y-%m-%d %H:%M")
    return v.strftime("%Y-%m-%d")


def _date_map(ws):
    """시트당 1회 빌드되는 {(row, col): 표시값} 맵 (날짜 + 퍼센트 표시값 복원).

    MergedCell(병합 비-앵커)은 is_date 속성이 없어 AttributeError 를 내므로
    isinstance 로 먼저 걸러낸다. is_date 만으로는 time(0,0) 등 datetime/date 가 아닌
    셀을 '1900-01-01' 로 오포맷할 수 있어 isinstance(value,(datetime,date)) 가드 필수.

    퍼센트: 서식 '0.0%' 셀의 저장값(0.3995…)을 엑셀 표시값(40.0%)으로 복원해 합류.
    bool 은 isinstance(True, int) 함정을 피하려 명시 제외. matrix desc 경로(L664)와
    orow=None 행은 오버라이드 미적용 — 날짜도 동일한 알려진 한계이며 코퍼스 0건.
    """
    out = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):   # 비-앵커 → is_date 접근 자체 회피
                continue
            if cell.is_date and isinstance(cell.value, (datetime, date)):
                out[(cell.row, cell.column)] = _fmt_dt(cell.value)
            elif (_is_pct_format(cell.number_format)
                  and isinstance(cell.value, (int, float))
                  and not isinstance(cell.value, bool)):
                d = _pct_decimals(cell.number_format)
                out[(cell.row, cell.column)] = f"{cell.value * 100:.{d}f}%"
    # 병합범위: 앵커(top-left)에만 값/서식 → 범위 전체로 전파(_fill_vertical 로 끌려온 자식셀도 잡힘)
    for rng in ws.merged_cells.ranges:
        a = (rng.min_row, rng.min_col)
        if a in out:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    out[(r, c)] = out[a]
    return out


def _strikethrough_cells(ws):
    """원본 XLSX의 취소선 좌표. kordoc에서 소실된 스타일을 k2o 좌표로 복원한다."""
    out = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None and bool(getattr(cell.font, "strike", False)):
                out.add((cell.row, cell.column))
    # 병합 anchor의 장식은 논리 셀 전체에 적용한다.
    for rng in ws.merged_cells.ranges:
        if (rng.min_row, rng.min_col) in out:
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    out.add((row, col))
    return out


def _original_cell_text(odate, ostrike, orow, col, kordoc_text):
    value = odate.get((orow, col), clean_val(kordoc_text))
    return mark_strikethrough(value, bool(orow and (orow, col) in ostrike))


# ─── kordoc .md → grid ───────────────────────────────
class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rawrows: List[List[Tuple[str, int, int, str]]] = []
        self._row = None
        self._buf: List[str] = []
        self._cs = self._rs = 1
        self._tag = "td"

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._buf = []
            self._cs = int(a.get("colspan", 1) or 1)
            self._rs = int(a.get("rowspan", 1) or 1)
            self._tag = tag
        elif tag == "br":
            self._buf.append("\n")

    def handle_data(self, d):
        if self._row is not None:
            self._buf.append(d)

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._row is not None:
            self._row.append(("".join(self._buf).strip(), self._cs, self._rs, self._tag))
        elif tag == "tr" and self._row is not None:
            self.rawrows.append(self._row)
            self._row = None


def _expand(rawrows):
    anchors, covered, occupied = {}, {}, set()
    ncols = 0
    for r, row in enumerate(rawrows):
        col = 1
        for (text, cs, rs, tag) in row:
            while (r, col) in occupied:
                col += 1
            for dr in range(rs):
                for dc in range(cs):
                    occupied.add((r + dr, col + dc))
                    covered[(r + dr, col + dc)] = (r, col)
            anchors[(r, col)] = (text, cs, rs, tag)
            col += cs
            ncols = max(ncols, col - 1)
    return anchors, covered, len(rawrows), ncols


def _parse_md_table(lines):
    rows = []
    for ln in lines:
        s = ln.strip()
        if not (s.startswith("|") and s.endswith("|")):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if cells and set("".join(cells).replace("-", "").replace(":", "").strip()) == set():
            continue
        rows.append([(c, 1, 1, "td") for c in cells])
    return rows


def _split_sheets(md):
    out, cur, body = [], None, []
    for ln in md.splitlines():
        m = re.match(r"^##\s+(.*)$", ln)
        if m:
            if cur is not None:
                out.append((cur, body))
            cur, body = m.group(1).strip(), []
        elif cur is not None:
            body.append(ln)
    if cur is not None:
        out.append((cur, body))
    return out


def _grid(body_lines):
    text = "\n".join(body_lines)
    if "<table" in text:
        tp = _TableParser()
        tp.feed(text)
        return _expand(tp.rawrows)
    return _expand(_parse_md_table(body_lines))


# ─── 원본 좌표 (k2o) ─────────────────────────────────
def _original_eff(ws):
    covered = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                covered[(r, c)] = (rng.min_row, rng.min_col)
    eff = {}
    for row in ws.iter_rows():
        for cell in row:
            eff[(cell.row, cell.column)] = cell.value

    def effective(r, c):
        a = covered.get((r, c), (r, c))
        return eff.get(a)
    return effective


def _align_rows(anchors, covered, nrows, ncols, ws):
    oeff = _original_eff(ws)

    def keff(r, c):
        a = covered.get((r, c))
        return anchors[a][0] if a in anchors else None

    k_sig = ["".join(norm(keff(r, c)) for c in range(1, ncols + 1)) for r in range(nrows)]
    o_sig = ["".join(norm(oeff(o, c)) for c in range(1, ncols + 1)) for o in range(1, ws.max_row + 1)]
    k2o = {}
    for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(None, k_sig, o_sig, autojunk=False).get_opcodes():
        if tag in ("equal", "replace"):
            for di in range(min(i2 - i1, j2 - j1)):
                k2o[i1 + di] = j1 + di + 1
    return k2o


# ─── 섹션 분할 ───────────────────────────────────────
def _row_cells(anchors, nrows):
    rc: Dict[int, List] = {}
    for (r, c), (text, cs, rs, tag) in anchors.items():
        rc.setdefault(r, []).append((c, text, cs, rs, tag))
    for r in rc:
        rc[r].sort()
    return rc


def _is_marker_row(nonempty) -> bool:
    return any(is_marker(t) for (_c, t, _cs, _rs, _tag) in nonempty)


def _multiheader_maps(group_cells, detail_cells):
    """2단 헤더 → (leaf_map{col:상세라벨}, group_map{col:그룹라벨}).

    group_cells(상위 스팬 헤더)와 detail_cells(아래 상세 헤더)를 합쳐 각 열의 라벨을 만든다.
    leaf 는 matrix 라벨(예: 팀장)에 그대로 쓰고, group 은 flat 행에서 접두사(업무현황_시스템구분)로 쓴다.
    """
    group_cov: Dict[int, str] = {}
    for (c, t, cs, _rs, _tag) in group_cells:
        lab = t.replace("\n", "").strip()
        if not lab:
            continue
        for col in range(c, c + max(1, cs)):
            group_cov[col] = lab
    leaf: Dict[int, str] = {}
    for (c, t, _cs, _rs, _tag) in detail_cells:
        lab = t.replace("\n", "").strip()
        if lab:
            leaf[c] = lab
    for col, g in group_cov.items():                 # 상세 라벨 없는 열은 그룹 라벨로 대체
        leaf.setdefault(col, g)
    group_map = {col: g for col, g in group_cov.items() if leaf.get(col) != g}
    return leaf, group_map


def _groupish_label(lab: str) -> bool:
    """pseudo-colspan run 을 이룰 수 있는 라벨: 문자(한글/영문) 포함 2자 이상.
    데이터 행의 반복 플레이스홀더('-', 'X', 숫자, 마커)가 그룹행으로 오인되는 것을 막는다."""
    return (len(lab) >= 2 and not is_marker(lab)
            and bool(re.search(r"[^\W\d_]", lab)))


def _group_spans(group_ne):
    """그룹행 후보의 유효 스팬 [(lo, hi)]: 실제 colspan≥2 + 연속 동일 텍스트 run(pseudo-colspan).

    접근제어_조사처럼 그룹 헤더를 병합 대신 같은 라벨 반복(최준영|최준영|최준영)으로 표기한
    시트를 위해, 열이 인접하고 텍스트가 같은 셀들을 하나의 스팬으로 합친다.
    병합(colspan≥2) 스팬은 라벨 무관 인정, 반복 run 은 _groupish_label 라벨만 인정."""
    spans = []
    prev = None                                      # (lo, hi, label, merged)
    for (c, t, cs, _rs, _tag) in group_ne:
        lab = t.replace("\n", "").strip()
        hi = c + max(1, cs) - 1
        if (prev is not None and lab == prev[2] and c == prev[1] + 1
                and _groupish_label(lab)):
            prev = (prev[0], hi, lab, True)
            continue
        if prev is not None:
            spans.append(prev)
        prev = (c, hi, lab, cs >= 2)
    if prev is not None:
        spans.append(prev)
    return [(lo, hi) for (lo, hi, _lab, ok) in spans if hi > lo and ok]


def _looks_multiheader(group_ne, detail_ne) -> bool:
    """group_ne(헤더 후보행)에 스팬(병합 colspan 또는 반복 텍스트 run)이 있고 detail_ne(다음
    비어있지 않은 행)이 그 아래 상세 헤더로 보이면 True. 단층 헤더(위임전결: 전결사항 colspan4
    + 아래는 ○ 데이터) 오탐을 막기 위해 엄격히 판정한다."""
    if not detail_ne:
        return False
    spans = _group_spans(group_ne)
    if not spans:
        return False
    if _is_marker_row(detail_ne):                    # 다음 행이 마커(○ 등) 데이터면 헤더 아님
        return False
    # 상세행은 '세분화되는 스팬 그룹' 수보다 많아야 한다.
    # rowspan≥2 leaf열(WBSID 등, 아래로 뻗어 detail에 안 나타남)은 세분화 대상이 아니라 제외 —
    # len(detail)==len(group) 여도(rowspan leaf + 스팬 혼합) 다단헤더를 놓치지 않게 한다.
    if len(detail_ne) <= len(spans):                 # 상세 헤더는 스팬을 더 잘게 나눈다
        return False
    single = sum(1 for (_c, _t, cs, _rs, _tag) in detail_ne if cs == 1)
    if single / max(1, len(detail_ne)) < 0.7:        # 상세행은 대부분 1칸
        return False
    span_cols = set()
    for (lo, hi) in spans:
        span_cols.update(range(lo, hi + 1))
    detail_cols = {c for (c, _t, _cs, _rs, _tag) in detail_ne}
    return bool(detail_cols & span_cols)             # 상세행이 스팬 구간을 채워야 함


def _longest_contiguous(cols) -> int:
    if not cols:
        return 0
    best = run = 1
    for a, b in zip(cols, cols[1:]):
        run = run + 1 if b == a + 1 else 1
        best = max(best, run)
    return best


def _fill_vertical(r, cells, anchors, covered, band):
    """세로병합(rowspan) 값을 하위 데이터 행에 채운다. band 안의 열만, 위에서 내려온 것(ar<r)만.
    가로 colspan 전파(ar==r)는 제외."""
    if band is None:
        return cells
    present = {c for (c, *_x) in cells}
    out = list(cells)
    for col in range(band[0], band[1] + 1):
        if col in present:
            continue
        a = covered.get((r, col))
        if a is None or a not in anchors:
            continue
        ar, _ac = a
        if ar < r:
            text, _cs, _rs, tag = anchors[a]
            if text.strip():
                out.append((col, text, 1, 1, tag))
    out.sort()
    return out


def _header_band(cells, max_gap=2):
    """헤더 행의 colspan-반영 커버 열들 중 최장(준)연속 구간 (lo, hi). 갭 너머 범례셀 배제용.
    라벨 없는 헤더 열 1~2개(접근제어_조사 '불필요' 열 등)로 반쪽이 잘리지 않게 max_gap 까지 허용."""
    covered = set()
    for (c, t, cs, _rs, _tag) in cells:
        if not t.strip():
            continue
        for col in range(c, c + max(1, cs)):
            covered.add(col)
    if not covered:
        return None
    cols = sorted(covered)
    segs = []
    lo = prev = cols[0]
    for c in cols[1:]:
        if c - prev - 1 > max_gap:
            segs.append((lo, prev))
            lo = c
        prev = c
    segs.append((lo, prev))
    return max(segs, key=lambda s: s[1] - s[0])


def _is_body_row(cells) -> bool:
    """항목번호(1./가./1.1.1) 또는 마커(○ 등)가 있으면 본문 시작 신호."""
    for (_c, t, _cs, _rs, _tag) in cells:
        s = t.strip()
        if not s:
            continue
        if is_marker(s) or infer_numbering_level(s) is not None:
            return True
    return False


def _row_header_score(cells, ncols) -> float:
    """스타일-free 헤더 점수. 텍스트/짧은라벨/연속열밴드/밀도 가산, 숫자/수식오류/번호/마커 감점."""
    filled = len(cells)
    if filled == 0:
        return -1e9
    nonnum = short = numeric = ferr = numbering = marker = 0
    for (_c, t, _cs, _rs, _tag) in cells:
        txt = t.replace("\n", " ").strip()
        compact = re.sub(r"\s+", "", txt)
        is_num = bool(_NUMERIC_RE.match(compact))
        is_ferr = bool(_FORMULA_ERR_RE.search(txt))
        is_mark = is_marker(txt)
        is_numb = infer_numbering_level(txt) is not None
        if is_num:
            numeric += 1
        if is_ferr:
            ferr += 1
        if is_mark:
            marker += 1
        if is_numb:
            numbering += 1
        if not (is_num or is_ferr):
            nonnum += 1
        if len(txt) <= 14:
            short += 1
    f = float(filled)
    cols = sorted(c for (c, _t, _cs, _rs, _tag) in cells)
    contig = _longest_contiguous(cols) / max(1, ncols)
    density = filled / max(1, (cols[-1] - cols[0] + 1))
    return (
        0.8 * (nonnum / f) + 0.5 * (short / f) + 0.6 * contig + 0.4 * density
        - 0.8 * (numeric / f) - 1.0 * (ferr / f) - 1.5 * (numbering / f) - 1.5 * (marker / f)
    )


def _pick_header(window):
    """window: 연속 헤더 후보 [(idx, r, cells), …] (single_full/본문행 제외, len>=2).
    반환 (header_idx, group_idx|None). 다단헤더(그룹행+상세행) 쌍 우선, 없으면 점수 최고행."""
    if not window:
        return None, None
    for k in range(len(window) - 1):
        _gi, _gr, gcells = window[k]
        di, _dr, dcells = window[k + 1]
        if _looks_multiheader(gcells, dcells):
            return di, window[k][0]
    ncols_eff = max((max(c for (c, *_x) in cells) for (_i, _r, cells) in window), default=1)
    best = max(window, key=lambda w: _row_header_score(w[2], ncols_eff))
    return best[0], None


def _segment(anchors, covered, nrows, ncols):
    rc = _row_cells(anchors, nrows)
    seq = []
    for r in range(nrows):
        nonempty = [(c, t, cs, rs, tag) for (c, t, cs, rs, tag) in rc.get(r, []) if t.strip()]
        if nonempty:
            seq.append((r, nonempty))
    # 다중셀(표 본문) 행들의 최소 시작열 — 소제목 판정의 ★주신호. _row_cells 가 정렬 반환하나
    # cells[0][0] 의존을 피해 min(cc for …) 로 안전하게 계산.
    body_min_col = min((min(cc for (cc, _t, _cs, _rs, _tag) in cells)
                        for (_r, cells) in seq if len(cells) >= 2), default=None)

    def _is_single_left(cells) -> bool:
        """부분병합 소제목: 다중셀 행 최소시작열 왼쪽에서 시작하는 병합 단일셀(전폭 아님)로,
        번호매김이 있거나 짧은 라벨이면 소제목. single_full(전폭)은 별개 소관이라 배제."""
        if len(cells) != 1 or body_min_col is None:
            return False
        c, t, cs = cells[0][0], cells[0][1], cells[0][2]
        if cs >= max(2, ncols - 1):          # 전폭이면 기존 single_full 소관
            return False
        if cs < 2:                           # 병합셀만(미병합 낱셀 배제)
            return False
        if c >= body_min_col:                # ★주신호: 본문 최소시작열보다 왼쪽이어야 함
            return False
        return (infer_numbering_level(t) is not None
                or len(t.strip()) <= _SUBTITLE_MAX_LEN)

    sheet_title = None
    sections: List[Dict[str, Any]] = []
    cur = None
    banner = None
    i = 0
    while i < len(seq):
        r, nonempty = seq[i]
        single_full = len(nonempty) == 1 and nonempty[0][2] >= max(2, ncols - 1)
        single_left = _is_single_left(nonempty)
        is_th = any(tag == "th" for (_, _, _, _, tag) in nonempty)
        # ① 순수장식 행(구분선 '----' 배너 등)은 시트제목/섹션조인 전에 skip — 제목·kv키·경로 오염 방지.
        #    (스텝 0 실측: 코퍼스 전역에서 R&R 구분선만 해당, 순수장식 데이터행 0건 → 데이터 소실 없음.)
        if all(_DECOR_RE.match(t.strip()) for (_c, t, _cs, _rs, _tag) in nonempty):
            i += 1
            continue
        if sheet_title is None and is_th and single_full:
            sheet_title = clean_title(nonempty[0][1])
            i += 1
            continue
        if single_full:
            title = _desecorate_title(clean_title(nonempty[0][1]))  # ② 조인 잔여 장식 2차 정화
            if not title:
                i += 1
                continue
            if cur is not None and cur["header"] is None and not cur["rows"]:
                cur["title"] = (cur["title"] + " / " + title) if cur["title"] else title
            else:
                cur = {"title": title, "header": None, "header_group": {}, "header_band": None, "rows": []}
                sections.append(cur)
            banner = cur["title"]          # 섹션 배너 갱신 — 이후 소제목을 'banner / sub' 로 조인
            i += 1
            continue
        if single_left:
            # 부분병합 소제목 → 배너 아래 하위 섹션(새 섹션 시작 → 반복 헤더의 표 경계).
            sub = _desecorate_title(clean_title(nonempty[0][1]))
            if sub:
                cur = {"title": (f"{banner} / {sub}" if banner else sub),
                       "header": None, "header_group": {}, "header_band": None, "rows": []}
                sections.append(cur)
            i += 1
            continue
        if cur is None:
            cur = {"title": None, "header": None, "header_group": {}, "header_band": None, "rows": []}
            sections.append(cur)
        if cur["header"] is None and len(nonempty) >= 2 and not _is_body_row(nonempty):
            # 헤더 후보 윈도우: i 부터 본문시작/제목 전까지(연속), 최대 _MAX_HEADER_SCAN
            window = []
            j = i
            while j < len(seq) and len(window) < _MAX_HEADER_SCAN:
                rj, cellsj = seq[j]
                sfj = len(cellsj) == 1 and cellsj[0][2] >= max(2, ncols - 1)
                if sfj or _is_single_left(cellsj) or _is_body_row(cellsj):
                    break
                if len(cellsj) >= 2:
                    window.append((j, rj, cellsj))
                j += 1
            hidx, gidx = _pick_header(window)
            if hidx is not None:
                if cur["rows"]:
                    # 헤더 = 표 시작 경계. 헤더 위에 이미 쌓인 행(부분병합 제목·메모·산문)은
                    # 표 소속이 아니므로 header-less 섹션으로 분리 → no_header emit 경로
                    # (컬럼레터/스팬 키, 값 있는 셀만)로 흘려 오라벨·빈값 청크를 차단한다.
                    # 정상 표(헤더 최상단)는 rows 가 빈 상태라 이 분기를 타지 않는다.
                    cur = {"title": cur["title"], "header": None, "header_group": {},
                           "header_band": None, "rows": []}
                    sections.append(cur)
                hcells = seq[hidx][1]
                if gidx is not None:
                    # 다단헤더: band 는 그룹행(rowspan leaf열 포함) + 상세행 합집합으로.
                    # 상세행만 보면 rowspan leaf열(WBSID 등)이 band 밖으로 밀려 drop됨.
                    cur["header_band"] = _header_band(seq[gidx][1] + hcells)
                    leaf, gmap = _multiheader_maps(seq[gidx][1], hcells)
                    cur["header"] = leaf
                    cur["header_group"] = gmap
                else:
                    cur["header_band"] = _header_band(hcells)
                    cur["header"] = {c: t.replace("\n", "").strip() for (c, t, cs, rs, tag) in hcells}
                i = hidx + 1                          # 헤더 위 메타/제목/범례 행은 drop
                continue
        filled = _fill_vertical(r, nonempty, anchors, covered, cur.get("header_band"))
        cur["rows"].append((r, filled))
        i += 1
    return sheet_title, sections


# ─── 청크 빌드 ───────────────────────────────────────
def _kw(*texts):
    toks, seen, out = [], set(), []
    for t in texts:
        for part in re.split(r"[\s/,·>()\[\]:;]+", str(t)):
            p = part.strip()
            if len(p) >= 2 and not is_marker(p):
                toks.append(p)
    for t in toks:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:30]


def _quality(conf, review, warns):
    return {"confidence": round(conf, 2), "review_required": bool(review),
            "warnings": warns, "parser_version": PARSER_VERSION}


def _build(doc, sheet, sheet_title, anchors, covered, nrows, ncols, k2o, odate, ostrike,
           content_cap=2000, embed_cap=3000):
    title = sheet_title or sheet
    _st, sections = _segment(anchors, covered, nrows, ncols)
    chunks: List[Dict[str, Any]] = []
    for sec in sections:
        sec_title = sec["title"]
        header = sec["header"]
        header_group = sec.get("header_group") or {}
        header_band = sec.get("header_band")
        rows = sec["rows"]
        if not rows:
            continue
        path = [title] + ([sec_title] if sec_title else [])
        no_header = header is None

        def _in_band(c):
            return header_band is None or (header_band[0] <= c <= header_band[1])

        # ── matrix_fact 마커 열 게이트: 섹션당 1-pass 열 집계(행 순회 전) ──
        # 마커 셀은 그 열이 '마커 전용 열'일 때만 matrix 지표. col_nonblank 는 마커 포함(분모)
        # — 순수 마커 열이 0/0 으로 전멸하지 않도록. 집계 소스는 _fill_vertical 적용 후 sec["rows"].
        # header-less 섹션에서도 text_cells 필터가 marker_col 을 참조하므로 keymap 블록 밖에 둔다.
        col_nonblank: Dict[int, int] = {}
        col_marker: Dict[int, int] = {}
        for (_r, _cells) in rows:
            for (_c, _t, _cs, _rs, _tag) in _cells:
                if not _in_band(_c):
                    continue
                _s = _t.strip()
                if not _s:
                    continue
                col_nonblank[_c] = col_nonblank.get(_c, 0) + 1
                if is_marker(_s):
                    col_marker[_c] = col_marker.get(_c, 0) + 1

        def marker_col(c):
            n = col_nonblank.get(c, 0)
            return n > 0 and col_marker.get(c, 0) / n >= _MARKER_COL_MIN

        def _label(c):
            leaf = header.get(c) if header else None
            if not leaf:
                return get_column_letter(c)
            g = header_group.get(c)
            return f"{g}_{leaf}" if g and g != leaf else leaf

        # 중복 라벨(담당자·도입일 등) → 2회차부터 컬럼레터 접미사로 고유화. 섹션당 1회 고정.
        keymap: Dict[int, str] = {}
        if header:
            seen: set = set()
            for c in sorted(header):
                base = _label(c)
                keymap[c] = base if base not in seen else f"{base}({get_column_letter(c)})"
                seen.add(base)

        def _key(c):
            return keymap.get(c) or _label(c)

        def emit_row(cells, orow, rng, source, review, conf, warns):
            present = {c: _original_cell_text(odate, ostrike, orow, c, t)
                       for (c, t, cs, rs, tag) in cells if t.strip() and _in_band(c)}
            if header:
                # 헤더 컬럼 전체(빈칸 포함) + 밴드 내 추가값 → 빈 헤더도 헤더="" 로 보존
                cols = sorted(c for c in (set(header) | set(present)) if _in_band(c))
                fields = {_key(c): present.get(c, "") for c in cols}
            else:
                # header-less: 컬럼레터 대신 의미 키(제목/주석) + A1 좌표 키.
                # 판정 순서 고정: note → title → 좌표. 값 v(=present[c], clean_val/odate 처리)로 판정.
                span = {c: cs for (c, t, cs, rs, tag) in cells if t.strip()}
                nvals = len(present)
                def _colkey(c):
                    cs = span.get(c, 1)
                    if orow:  # 좌표 해소됨 → A1(단일 'D4' / 병합 'D2:G2')
                        return range_a1(orow, c, orow, c + cs - 1)
                    # coord_unresolved 폴백: 레터/스팬 레터
                    return get_column_letter(c) if cs <= 1 else \
                        f"{get_column_letter(c)}-{get_column_letter(c + cs - 1)}"
                def _semkey(c, v):
                    v = str(v)
                    if not is_marker(v) and (v.startswith(_HL_NOTE_PREFIXES) or _BYULPYO_RE.match(v)):
                        return "주석"
                    if nvals == 1 and infer_numbering_level(v) is None and len(v.strip()) <= _SUBTITLE_MAX_LEN:
                        return "제목"
                    return _colkey(c)
                fields = {}
                for c, v in present.items():
                    k = _semkey(c, v)
                    if k in fields:  # 중복 키(주석 2개 등) → keymap 관례대로 컬럼레터 접미 고유화
                        k = f"{k}({get_column_letter(c)})"
                    fields[k] = v
            if not fields or not any(str(v).strip() for v in fields.values()):
                return  # 키만 있고 값 전부 공란(밴드밖 행 유래) → 빈값 청크 차단
            fld_txt = ", ".join(f"{k}: {v}" for k, v in fields.items())
            content_frame = build_content_frame(sheet, title, sec_title)
            content = f"{content_frame} 항목: {fld_txt}"
            core = (f"title: {title}; path: {' > '.join(path)}; "
                    + "; ".join(f"{k}: {v}" for k, v in fields.items()) + f" -- {sheet} [{rng}]")
            chunks.append({
                "id": f"{doc}::{sheet}::{rng}::row", "source_file": doc, "sheet": sheet, "range": rng,
                "chunk_type": "table_row", "region_type": ("unknown_table" if no_header else "flat_table"),
                "title": title, "path": path, "fields": fields,
                "facts": [{"predicate": k, "value": v} for k, v in fields.items()],
                "content_text": content[:content_cap], "keywords": _kw(title, sec_title, *fields.values()),
                "source": source,
                "metadata": {"workbook_title": title, "section": sec_title,
                             "content_frame": content_frame,
                             "core_text": core[:embed_cap], "embedding_text": core[:embed_cap]},
                "quality": _quality(conf, review, warns),
            })

        for (r, cells) in rows:
            orow = k2o.get(r)
            cols = [c for (c, t, cs, rs, tag) in cells if t.strip()]
            if not cols:
                continue
            c1, c2 = min(cols), max(cols)
            rng = f"{get_column_letter(c1)}{orow}:{get_column_letter(c2)}{orow}" if orow else None
            source = {"file": doc, "sheet": sheet, "range": rng, "start_row": orow,
                      "end_row": orow, "start_col": c1, "end_col": c2}
            warns = (["header_not_detected"] if no_header else []) + ([] if orow else ["coord_unresolved"])
            review = no_header or (orow is None)
            conf = 0.86 if (not no_header and orow) else 0.6

            marker_cells = [(c, t.strip()) for (c, t, cs, rs, tag) in cells
                            if is_marker(t) and _in_band(c) and marker_col(c)
                            and not (orow and (orow, c) in ostrike)]
            # 게이트 탈락 마커(값 열에 우발한 △ 등)는 값으로 desc/fields 에 보존(must-fix①).
            text_cells = [(c, t) for (c, t, cs, rs, tag) in cells
                          if t.strip()
                          and not (
                              is_marker(t) and marker_col(c)
                              and not (orow and (orow, c) in ostrike)
                          )
                          and _in_band(c)]
            if header and marker_cells:
                primary_label = keymap.get(min(header)) if header else None
                desc = {
                    (keymap.get(c) or primary_label or get_column_letter(c)):
                    _original_cell_text(odate, ostrike, orow, c, t)
                    for (c, t) in text_cells
                }
                row_label = " > ".join(desc.values()) if desc else f"row{orow}"
                buckets: Dict[str, List[str]] = {}
                for (c, t) in marker_cells:
                    bk = MARK_BUCKET.get(t, "해당")
                    buckets.setdefault(bk, []).append(_key(c))
                fields = dict(desc)
                for bk, cl in buckets.items():
                    fields[bk] = ", ".join(cl)
                grp = "; ".join(f"{bk}: {', '.join(cl)}" for bk, cl in buckets.items())
                content_frame = build_content_frame(sheet, title, sec_title, row_label)
                content = f"{content_frame} 항목: {grp}"
                core = (f"title: {title}; path: {' > '.join(path)}; "
                        + "; ".join(f"{k}: {v}" for k, v in fields.items()) + f" -- {sheet} [{rng}]")
                facts = [{"subject": row_label, "predicate": bk, "object": col}
                         for bk, cl in buckets.items() for col in cl]
                chunks.append({
                    "id": f"{doc}::{sheet}::{rng}::matrix", "source_file": doc, "sheet": sheet, "range": rng,
                    "chunk_type": "matrix_fact", "region_type": "matrix_table", "title": title,
                    "path": path + [row_label], "fields": fields, "facts": facts,
                    "content_text": content[:content_cap],
                    "keywords": _kw(title, sec_title, row_label, *[v for cl in buckets.values() for v in cl]),
                    "source": source,
                    "metadata": {"workbook_title": title, "section": sec_title,
                                 "content_frame": content_frame,
                                 "core_text": core[:embed_cap], "embedding_text": core[:embed_cap]},
                    "quality": _quality(conf, review, warns),
                })
            else:
                emit_row(cells, orow, rng, source, review, conf, warns)
    return chunks


# ─── md 확보 ─────────────────────────────────────────
def _resolve_md(input_path: Path, config: ParserConfig) -> str:
    stem = input_path.stem
    if config.kordoc_md_path and Path(config.kordoc_md_path).exists():
        return Path(config.kordoc_md_path).read_text(encoding="utf-8")
    if config.kordoc_md_dir:
        p = Path(config.kordoc_md_dir) / f"{stem}.md"
        if p.exists():
            return p.read_text(encoding="utf-8")
    # 자동 생성 (Node 필요)
    if config.kordoc_bin:
        out_dir = Path(config.kordoc_md_out or tempfile.gettempdir())
        out_dir.mkdir(parents=True, exist_ok=True)
        md_path = out_dir / f"{stem}.md"
        # kordoc CLI: `kordoc <file> -o <out.md> --silent`. kordoc_bin 은 멀티워드 허용
        # (예: "node /path/dist/cli.js").
        cmd = shlex.split(config.kordoc_bin) + [str(input_path), "-o", str(md_path), "--silent"]
        proc = subprocess.run(cmd, capture_output=True, text=True)
        if proc.returncode != 0 or not md_path.exists():
            raise BackendError(
                f"kordoc 자동생성 실패 (bin={config.kordoc_bin}): {proc.stderr[:300] or proc.stdout[:300]}")
        return md_path.read_text(encoding="utf-8")
    raise BackendError(
        f"kordoc backend: '{stem}.md' 를 찾을 수 없습니다. "
        f"--kordoc-md/--kordoc-md-dir 로 지정하거나 --kordoc-bin 으로 자동생성하세요.")


class KordocBackend:
    def parse(self, input_path: Path, config: ParserConfig) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        md = _resolve_md(input_path, config)
        doc = config.document_title or input_path.stem
        wb = load_workbook(input_path, data_only=True)
        ksheets = dict(_split_sheets(md))
        all_chunks: List[Dict[str, Any]] = []
        sheets_done = 0
        for sheet in wb.sheetnames:
            body = ksheets.get(sheet)
            if body is None:
                continue
            anchors, covered, nrows, kncols = _grid(body)
            ws = wb[sheet]
            acols = max(kncols, ws.max_column)
            k2o = _align_rows(anchors, covered, nrows, acols, ws)
            odate = _date_map(ws)
            ostrike = _strikethrough_cells(ws)
            sheet_title, _secs = _segment(anchors, covered, nrows, kncols)
            all_chunks.extend(_build(doc, sheet, sheet_title, anchors, covered, nrows, kncols,
                                     k2o, odate, ostrike,
                                     content_cap=config.row_content_max_chars,
                                     embed_cap=config.row_embedding_max_chars))
            sheets_done += 1

        # CGH 계층 병합 — self-gating(번호 spine 없으면 무변화). 모든 내부노드가
        # 직속 자식 아웃라인을 품는 hierarchy_node 요약청크를 발행한다.
        from ..chunking.hierarchy_tree import merge_hierarchy_rows
        all_chunks = merge_hierarchy_rows(all_chunks, max_chars=config.numbering_merge_max_chars)

        # 공통 형제 묶음(사용자 승인 b) — section 아래 같은 (sheet,path) 연속 run 을
        # sibling_rule 로 대체. CGH 가 spine 병합한 시트는 잔여 run 이 게이트/상한에 걸려
        # 대부분 무변화. sibling_rule_max_chars=0 이면 비활성.
        from ..chunking.sibling_rule import build_kordoc_siblings, _reorder_kordoc_siblings
        siblings, replaced = build_kordoc_siblings(all_chunks, config.sibling_rule_max_chars)
        if siblings:
            # 문서 순서 보존(사용자 요청): sibling 을 첫 멤버 원본 위치에 삽입(말미 append 폐기).
            all_chunks = _reorder_kordoc_siblings(all_chunks, siblings, replaced)

        ct = Counter(c["chunk_type"] for c in all_chunks)
        rt = Counter(c["region_type"] for c in all_chunks)
        confs = [c["quality"]["confidence"] for c in all_chunks] or [0.0]
        stats = {
            "backend": "kordoc",
            "source_file": input_path.name,
            "sheet_count": len(wb.sheetnames),
            "sheets_parsed": sheets_done,
            "total_chunks": len(all_chunks),
            "chunk_type_counts": dict(ct),
            "region_type_counts": dict(rt),
            "coords": sum(1 for c in all_chunks if c["range"]),
            "confidence": {"avg": round(sum(confs) / len(confs), 4), "min": min(confs), "max": max(confs)},
            "review_required_count": sum(1 for c in all_chunks if c["quality"]["review_required"]),
        }
        return all_chunks, stats
