"""Header 탐지 (SoT §11) + 컬럼 역할 부여.

detect_headers(region, canvas, config):
- header_score 로 상단 연속 헤더 행 판정 (max_header_depth)
- 다단 헤더 flatten → 컬럼명 생성
  - 빈 헤더는 위/왼쪽 logical 로 보강 (병합은 logical_value 가 이미 처리)
  - matrix 계열 region 은 "전결권자" 같은 상위 그룹 헤더 대신 하위 라벨 우선
  - 이름 충돌 시 suffix
- region.header_rows / hierarchy_cols / matrix_cols / metadata_cols / body_rows 채움
- override 로 이미 채워진 값은 유지 (빈 이름만 보강)
"""

from __future__ import annotations

import copy
import re
from typing import Dict, List, Optional, Set, Tuple, TYPE_CHECKING

from openpyxl.utils import range_boundaries

from ..config import ParserConfig
from ..markerutil import is_marker_cell
from ..textutil import compact, infer_numbering_level, is_spine_column, one_line

if TYPE_CHECKING:
    from ..canvas.cell_node import CellNode
    from ..canvas.sheet_canvas import SheetCanvas
    from ..detection.region import Region

_HEADER_SCORE_MIN = 1.8
_STYLE_GATE_MIN = 0.5   # bold/배경색/가운데정렬 등 스타일 신호 최소값
_MAX_PRE_ROWS = 8       # 헤더 시작 전 스킵 가능한 희소 메타 행 수
_DATA_BREAK_MIN = 0.5   # numeric(날짜/숫자) 과반 = 데이터신호 → 시작된 헤더밴드 종료
_W_FILTER = 0.8         # autofilter 첫 행 가산점 (center/text 항과 동일 크기, override 아님)

# 콜론(:／：)으로 끝나는 라벨 셀 — 메타 key-value 행 감지용
_COLON_LABEL_RE = re.compile(r".+[:：]\s*$")

# 합의/수신/비고 류 — metadata_cols 로 분리할 헤더명 (compact 비교)
_METADATA_HEADER_TERMS = {"합의", "수신", "비고", "참고", "참조", "근거", "관련근거", "관련규정"}
# 항목/계층 컬럼 헤더로 보는 어휘 (compact 후 부분 일치)
_ITEM_HEADER_TERMS = ("사항", "항목", "구분", "내용", "업무", "분류", "제목", "품목", "품명")

# 헤더 탐지를 건너뛰는 region 유형 (key-value/주석/제목 류)
_HEADERLESS_TYPES = {"code_mapping_table", "form", "key_value_block", "note_block", "report_section"}

_NUMERIC_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?%?$")

# spine 후보에서 제외할 버전열 헤더 패턴 (kordoc _VERSION_KEY 이식 — chunking/hierarchy_tree.py).
# 분류 시점엔 헤더 미확정이라 min_depth=3 이 주 방어, 여긴 검출단계 이중 방어층.
_VERSION_KEY_RE = re.compile(r"버전|version|\bver\.?\b|개정|리비전|revision|rev\.?", re.I)

# 날짜 '문자열' (타입 date 가 아닌 텍스트 날짜 — "2026.05.04", "2026-5-4", "2026/05/04", "2026. 5. 4.")
# 사용자 스펙 §9-1: 날짜 문자열은 text 가 아니라 value. 4자리 연도 앵커로 버전문자열("1.2.3") 오탐 차단.
# 공백은 매치 전에 compact() 가 제거하므로 정규식에 \s 불필요.
_DATE_TEXT_RE = re.compile(r"^\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\.?$")


# --- 공용 헬퍼 -------------------------------------------------------------------

def _cell_text(cell: Optional["CellNode"]) -> str:
    if cell is None:
        return ""
    v = cell.logical_value or cell.normalized_value or cell.display_value
    if not v and cell.raw_value is not None:
        v = cell.raw_value
    return one_line(v)


def _is_numeric(cell: Optional["CellNode"], text: str) -> bool:
    # 날짜(로더는 datetime/date/time 을 전부 'date' 로 방출)도 값으로 인식 —
    # "시작일자: 2026-05-04" 같은 메타행이 텍스트 헤더로 오탐되는 것을 차단.
    if cell is not None and cell.data_type in ("int", "float", "date"):
        return True
    t = compact(text)
    return bool(_NUMERIC_RE.match(t)) or bool(_DATE_TEXT_RE.match(t))


def _is_meta_kv_row(canvas: "SheetCanvas", row: int, cols: List[int]) -> bool:
    """콜론라벨(`작성자:` 등) + 희소 채움 = 메타 key-value 행.

    A(날짜=value)가 날짜 있는 메타행을 해결하므로 B 는 날짜 없는 텍스트 메타행
    (`작성자: 홍길동` 등)을 방어한다. 정상 헤더 라벨(`구분:` 등)은 채움 비율이
    높아(>=0.5) 미발동 — 희소성 조건이 이중 가드.
    """
    n = len(cols)
    filled = colon = 0
    for c in cols:
        t = _cell_text(canvas.cells.get((row, c)) or canvas.get_cell(row, c))
        if not t:
            continue
        filled += 1
        if _COLON_LABEL_RE.match(t):
            colon += 1
    return filled > 0 and colon >= 1 and (filled / float(n)) < 0.5


def _is_axis_sequence_row(canvas: "SheetCanvas", row: int, cols: List[int]) -> bool:
    """행의 숫자 셀들이 step==1 단조 증가 수열이면 축(axis) 헤더 — 간트 연도/일자 타임라인.

    범위(v3 — 정밀도 우선, 검증관 실측 반영):
    - int/float 수열의 step == 1 만 축으로 인정 (연도 2024,2025,…(엑셀 float 2024.0 포함) / 일자 1,2,…).
      금액 100/200/300(step 100) 등 등차 데이터·역순(2026,2025,…)은 배제.
    - ★date 타입은 판정하지 않는다(무조건 False 경로): 로더가 datetime.time 도 'date' 로
      방출해 뺄셈이 TypeError(워크북 전체 파싱 중단 위험) + 골든 105 에 date-축 실례 0건
      + 주간 마일스톤 데이터행([task, 1/1, 1/8, 1/15]) 등간격 오탐면. 이득 0 인 위험 분기 삭제.
    - 수열 길이 >= 3, 그리고 행의 채워진 셀 중 수열 참여 셀이 과반.
    """
    nums = []
    filled = 0
    for c in cols:
        cell = canvas.cells.get((row, c))
        text = _cell_text(cell)
        if not text:
            continue
        filled += 1
        if cell is not None and cell.data_type in ("int", "float"):
            nums.append(float(cell.raw_value))
    if len(nums) < 3 or filled == 0 or len(nums) * 2 <= filled:
        return False
    diffs = {round(b - a, 6) for a, b in zip(nums, nums[1:])}
    return diffs == {1.0}   # step==1 단조 증가만 (연도/일자 축)


def _merge_span_cols(cell: "CellNode", canvas: "SheetCanvas") -> Tuple[int, int]:
    rng = cell.merge_range
    if not rng:
        for mr in canvas.merged_ranges:
            try:
                c0, r0, c1, r1 = range_boundaries(mr)
            except Exception:
                continue
            if r0 <= cell.row <= r1 and c0 <= cell.col <= c1:
                return int(c0), int(c1)
        return cell.col, cell.col
    try:
        c0, _r0, c1, _r1 = range_boundaries(rng)
        return int(c0), int(c1)
    except Exception:
        return cell.col, cell.col


def _is_merge_shadow_row(canvas: "SheetCanvas", row: int, header_rows: List[int]) -> bool:
    """row 의 셀 중 하나라도 header_rows 에서 시작하는 세로/블록 병합에 덮이면 True.

    스팬 부모헤더(전결권자 F1:G1) 아래의 무스타일·희소 leaf 행(부총장/처장)을
    헤더밴드로 편입하기 위한 신호. 세로병합(부서명 A1:A2 등)이 헤더행에서 시작해
    아래로 뻗으면 그 밑 행도 같은 헤더 구조의 일부다.
    """
    hset = set(header_rows)
    for mr in canvas.merged_ranges:
        try:
            _c0, r0, _c1, r1 = range_boundaries(mr)
        except Exception:
            continue
        if r0 in hset and r0 < row <= r1:
            return True
    return False


def _title_rows(region: "Region") -> Set[int]:
    """region.title_range 가 region 내부에 있으면 해당 행들을 반환."""
    if not region.title_range:
        return set()
    try:
        _c0, r0, _c1, r1 = range_boundaries(region.title_range)
    except Exception:
        return set()
    rows = {r for r in range(int(r0), int(r1) + 1) if region.min_row <= r <= region.max_row}
    return rows


def _autofilter_first_row(canvas: "SheetCanvas", region: "Region") -> Optional[int]:
    """auto_filter_ref 첫 행. 파싱 실패/리전 밖이면 None."""
    ref = getattr(canvas, "auto_filter_ref", None)
    if not ref:
        return None
    try:
        _c0, r0, _c1, _r1 = range_boundaries(ref)
    except Exception:
        return None
    r0 = int(r0)
    return r0 if region.min_row <= r0 <= region.max_row else None


# --- row metrics / header score (SoT §11.2) --------------------------------------

def _row_metrics(canvas: "SheetCanvas", row: int, cols: List[int]) -> Dict[str, float]:
    n = len(cols)
    filled = bold = center = fillc = border = nonnum = short = hmerge = marker = numbering = numeric = 0
    for c in cols:
        cell = canvas.cells.get((row, c))
        text = _cell_text(cell)
        if not text:
            continue
        filled += 1
        st = cell.style
        if st.bold:
            bold += 1
        if st.horizontal_alignment in ("center", "centerContinuous"):
            center += 1
        if st.fill_color:
            fillc += 1
        if st.border_top or st.border_bottom or st.border_left or st.border_right:
            border += 1
        if cell.merge_orientation in ("horizontal", "block"):
            hmerge += 1
        if is_marker_cell(cell, text):
            marker += 1
        if infer_numbering_level(text) is not None:
            numbering += 1
        if _is_numeric(cell, text):
            numeric += 1
        else:
            nonnum += 1
        if len(text) <= 14:
            short += 1
    if filled == 0:
        return {"filled": 0.0, "style_signal": 0.0}
    f = float(filled)
    m = {
        "filled": float(filled),
        "bold": bold / f,
        "center": center / f,
        "fillcolor": fillc / f,
        "border": border / f,
        "text": nonnum / f,
        "short": short / f,
        "hmerge": hmerge / f,
        "fill": filled / float(n),
        "marker_ratio": marker / f,
        "numbering_ratio": numbering / f,
        "numeric": numeric / f,
    }
    m["style_signal"] = m["bold"] + m["fillcolor"] + (0.5 if m["center"] >= 0.5 else 0.0)
    return m


def _header_score(m: Dict[str, float]) -> float:
    """SoT §11.2 — bold/fill/border/text/short/merge 가산, marker/번호/숫자 감점."""
    return (
        1.2 * m["bold"]
        + 0.8 * m["center"]
        + 0.6 * m["fillcolor"]
        + 0.4 * m["border"]
        + 0.8 * m["text"]
        + 0.5 * m["short"]
        + 0.4 * m["hmerge"]
        + 0.6 * m["fill"]
        - 1.5 * m["marker_ratio"]
        - 1.5 * m["numbering_ratio"]
        - 0.8 * m["numeric"]
    )


def _detect_header_rows(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    title_rows: Set[int],
    max_depth: int,
) -> Tuple[List[int], Set[int], float]:
    """상단에서 연속 header 행 탐지. (header_rows, pre_rows(헤더 위 메타행), best_score)"""
    pre: Set[int] = set()
    headers: List[int] = []
    best = 0.0

    # autofilter 첫 행(range_boundaries 의 상단 행) — style 동일 가중 가산점 대상.
    # 파싱 실패/미존재 시 None. filter 첫 행은 밴드 전체가 아니라 leaf 헤더행을 가리킨다.
    filter_header_row: Optional[int] = None
    ref = getattr(canvas, "auto_filter_ref", None)
    if ref:
        try:
            _c0, r0, _c1, _r1 = range_boundaries(ref)
            filter_header_row = int(r0)
        except Exception:
            filter_header_row = None

    r = region.min_row
    while r <= region.max_row:
        if r in title_rows:
            if headers:
                break
            pre.add(r)
            r += 1
            continue
        m = _row_metrics(canvas, r, cols)
        if m["filled"] == 0:
            if headers:
                break
            pre.add(r)
            r += 1
            continue
        # 메타 key-value 행은 밴드 시작 전에만 pre 로 divert (밴드 내 오발동 차단)
        if not headers and _is_meta_kv_row(canvas, r, cols):
            pre.add(r)
            r += 1
            continue
        if m.get("numbering_ratio", 0.0) > 0 or m.get("marker_ratio", 0.0) > 0:
            break  # 본문 시작 (항목 번호/마커 등장)
        if headers and m.get("numeric", 0.0) >= _DATA_BREAK_MIN:
            # 데이터신호(날짜/숫자 과반) = 시작된 헤더밴드의 종료.
            # 단, step==1 축 수열(간트 연도/일자 leaf)이면 밴드에 편입하고 종료한다.
            # 밴드 시작 전엔 미발동(메타행은 A/B 가 divert) — 게이트 없으면 WBS 헤더 전멸.
            if _is_axis_sequence_row(canvas, r, cols):
                headers.append(r)   # 축 leaf 헤더 — 밴드에 편입하고 밴드 종료
                break               # 축 행은 leaf: 그 아래는 본문 (shadow_cont 종료와 동일 원칙)
            break  # 기존 Change-C: 데이터신호 = 밴드 종료
        # autofilter 첫 행 가산: 데이터틱하지도(numeric>=0.5) 메타도 아닐 때만
        # (stale/전시트 ref 가 데이터·배너 행을 승격시키는 것 차단)
        is_filter = (
            filter_header_row is not None
            and r == filter_header_row
            and m.get("numeric", 0.0) < _DATA_BREAK_MIN
            and not _is_meta_kv_row(canvas, r, cols)
        )
        score = _header_score(m) + (_W_FILTER if is_filter else 0.0)
        eff_style = m["style_signal"] + (_STYLE_GATE_MIN if is_filter else 0.0)
        strong = (
            score >= _HEADER_SCORE_MIN
            and eff_style >= _STYLE_GATE_MIN
            and m["filled"] >= 2
        )
        # 스팬 부모헤더 아래 leaf 행: 무스타일·희소라 style gate 는 탈락하지만,
        # 헤더행에서 시작한 세로/블록 병합에 덮이면(=헤더밴드 연장) 헤더로 편입한다.
        # 마커/번호 행은 위에서 이미 break 되므로 본문행 오편입 위험 없음.
        shadow_cont = (
            not strong
            and bool(headers)
            and m["filled"] >= 2
            and score >= _HEADER_SCORE_MIN
            and _is_merge_shadow_row(canvas, r, headers)
        )
        if strong or shadow_cont:
            headers.append(r)
            best = max(best, score)
            # shadow leaf 행(스팬 부모 아래 sub-header)은 헤더밴드의 끝이다.
            # 그 아래는 본문이므로 더 진행하지 않는다(블록병합·스타일된 데이터
            # 카테고리 행이 strong 으로 오편입되는 것을 차단).
            if shadow_cont or len(headers) >= max_depth:
                break
        else:
            if headers:
                break
            # 헤더 시작 전 메타 행 ("<별표N>", "(개정 : ...)" 단독 셀 등) — 한도 내에서 스킵
            pre.add(r)
            if len(pre) >= _MAX_PRE_ROWS:
                break
        r += 1
    return headers, pre, best


# --- multi-row header flatten (SoT §11.4) -----------------------------------------

def _collect_header_parts(
    canvas: "SheetCanvas", cols: List[int], header_rows: List[int]
) -> Dict[int, List[str]]:
    parts_by_col: Dict[int, List[str]] = {}
    for c in cols:
        parts: List[str] = []
        for r in header_rows:
            v = _cell_text(canvas.cells.get((r, c)) or canvas.get_cell(r, c))
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        parts_by_col[c] = parts
    return parts_by_col


def _flatten_column_names(
    region: "Region", cols: List[int], parts_by_col: Dict[int, List[str]]
) -> Dict[int, str]:
    matrix_mode = region.region_type in ("matrix_table", "hierarchical_matrix")
    names: Dict[int, str] = {}
    used: Dict[str, int] = {}
    prev_name = ""
    for c in cols:
        parts = parts_by_col.get(c, [])
        if parts:
            # matrix 계열: 상위 그룹 헤더("전결권자")보다 하위 라벨("팀장") 우선
            name = parts[-1] if matrix_mode else "_".join(parts)
            prev_name = name
        else:
            # 빈 헤더 — 왼쪽 logical 보강 (왼쪽도 없으면 "")
            name = prev_name
        if name:
            used[name] = used.get(name, 0) + 1
            if used[name] > 1:
                name = f"{name}_{used[name]}"
        names[c] = name
    return names


# --- 컬럼 역할 부여 ----------------------------------------------------------------

def _absorb_label_cols(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    parts_by_col: Dict[int, List[str]],
    header_rows: List[int],
    first_named: int,
) -> List[int]:
    """first_named 부터 우측으로 '마커 희박한 좌측 라벨열'을 역할밴드 전까지 연속 흡수.

    다중 좌측 라벨열(구분+업무내용, 부서명+단위업무+세부업무)을 계층축으로 보존한다.
    중단 조건: 역할열(마커 밀집/순수마커) 또는 메타 헤더(비고/합의 등) 또는 빈 열.
    위치 하드코딩 없음 — 판정은 본문 마커밀도와 헤더 어휘로만.
    """
    body_start = (max(header_rows) + 1) if header_rows else region.min_row

    def _col_role(c: int) -> Tuple[int, bool]:
        """(filled, is_role). is_role = 마커 밀집(>=2) 또는 순수마커(mk==filled)."""
        filled = mk = 0
        for r in range(body_start, region.max_row + 1):
            cell = canvas.cells.get((r, c))
            if cell is None or cell.is_empty:
                continue
            t = _cell_text(cell)
            if not t:
                continue
            filled += 1
            if is_marker_cell(cell, t):
                mk += 1
        return filled, (filled > 0 and (mk >= 2 or mk == filled))

    colset = set(cols)
    absorbed: List[int] = [first_named]
    for c in cols:
        if c <= first_named:
            continue
        filled, isrole = _col_role(c)
        if isrole:
            # 진짜 역할밴드인가: 이 열과 다음 열이 모두 role(=≥2 연속). 자산목록처럼
            # 데이터열 사이에 흩어진 단일 boolean 마커열은 밴드가 아니므로 기각.
            nxt = c + 1
            _, nxt_role = _col_role(nxt) if nxt in colset else (0, False)
            if nxt_role and len(absorbed) >= 2:
                return absorbed
            return []
        # 역할 전에 메타(비고/합의 …) 또는 빈 열을 만나면 전결 매트릭스 레이아웃 아님 → 기각
        name = compact(parts_by_col[c][-1]) if parts_by_col.get(c) else ""
        if name in _METADATA_HEADER_TERMS or filled == 0:
            return []
        absorbed.append(c)
    return []  # 끝까지 연속 역할밴드를 만나지 못함 → 다중 라벨열 아님


def _detect_hierarchy_cols(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    parts_by_col: Dict[int, List[str]],
    header_rows: List[int],
) -> List[int]:
    colset = set(cols)
    first_named: Optional[int] = None
    for c in cols:
        if parts_by_col.get(c):
            first_named = c
            break

    if header_rows and first_named is not None:
        anchor = canvas.get_cell(header_rows[0], first_named)
        c0, c1 = _merge_span_cols(anchor, canvas)
        if c1 > c0:
            # "전 결 사 항" 처럼 항목 헤더가 가로(또는 블록) 병합으로 여러 컬럼을 차지
            return [c for c in range(c0, c1 + 1) if c in colset]
        # 매트릭스 계열: 다중 좌측 라벨열(구분+업무내용 등)을 역할밴드 전까지 흡수.
        if region.region_type in ("matrix_table", "hierarchical_matrix"):
            absorbed = _absorb_label_cols(
                region, canvas, cols, parts_by_col, header_rows, first_named
            )
            if len(absorbed) >= 2:
                return absorbed
        label = compact(parts_by_col[first_named][0])
        if any(term in label for term in _ITEM_HEADER_TERMS):
            return [first_named]

    # fallback: 왼쪽 컬럼들의 항목 번호 패턴 스캔
    body_start = (max(header_rows) + 1) if header_rows else region.min_row
    hier: List[int] = []
    for c in cols[: min(5, len(cols))]:
        texts: List[str] = []
        for r in range(body_start, region.max_row + 1):
            cell = canvas.cells.get((r, c))
            if cell is None or cell.is_empty:
                continue
            t = _cell_text(cell)
            if t and not is_marker_cell(cell, t):
                texts.append(t)
        if len(texts) < 3:
            continue
        hits = sum(1 for t in texts if infer_numbering_level(t) is not None)
        if hits / len(texts) >= 0.3:
            hier.append(c)
    if hier:
        return hier
    if region.region_type in ("matrix_table", "hierarchical_matrix") and first_named is not None:
        return [first_named]

    # fallback-of-fallback: hierarchical_table 인데 계층열 0개인 자기모순 해소 (Phase 2 분류
    # 세로병합 신호의 계층열 판정 대응물). 좌측 <=5열의 본문 세로병합 비율 — 가로/블록 병합
    # 셀은 행 배너성(그룹 요약행 통과)이라 분모에서 제외. 연속 prefix 만(우측 무병합 열에서 종료).
    if region.region_type == "hierarchical_table":
        hier2: List[int] = []
        for c in cols[: min(5, len(cols))]:
            vm = plain = 0
            for r in range(body_start, region.max_row + 1):
                cell = canvas.cells.get((r, c))
                if cell is None or cell.is_empty:
                    continue
                if not _cell_text(cell):
                    continue
                o = cell.merge_orientation
                if o == "vertical":
                    vm += 1
                elif o in ("horizontal", "block"):
                    pass
                else:
                    plain += 1
            if vm >= 3 and (vm + plain) > 0 and vm / (vm + plain) >= 0.5:
                hier2.append(c)
            elif hier2:
                break
        if hier2:
            return hier2

    # spine 게이트(v2): 좌측 ≤5열 body 값이 dotted-int 계층 spine(WBSID 1/1.1/1.1.1)이면 계층열.
    # 기존 신호 ①②③·merge fallback 보다 뒤(우선순위 유지). matrix 계열은 :494 조기 return 으로
    # 미도달(암묵 가드). +방어층: 헤더명이 버전열 패턴이면 제외(분류 시점 헤더 미확정 → depth>=3 이 주 방어).
    spine_hier: List[int] = []
    for c in cols[: min(5, len(cols))]:
        header_name = " ".join(parts_by_col.get(c) or [])
        if header_name and _VERSION_KEY_RE.search(header_name):
            continue
        texts: List[str] = []
        for r in range(body_start, region.max_row + 1):
            cell = canvas.cells.get((r, c))
            if cell is None or cell.is_empty:
                continue
            t = _cell_text(cell)
            if t and not is_marker_cell(cell, t):
                texts.append(t)
        if is_spine_column(texts):
            spine_hier.append(c)
    if spine_hier:
        return spine_hier
    return []


def _assign_column_roles(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    names: Dict[int, str],
    parts_by_col: Dict[int, List[str]],
    header_rows: List[int],
) -> None:
    if not region.hierarchy_cols:
        region.hierarchy_cols = _detect_hierarchy_cols(region, canvas, cols, parts_by_col, header_rows)
    hier = set(region.hierarchy_cols)

    if region.metadata_cols:
        for c, nm in list(region.metadata_cols.items()):
            if not nm:
                region.metadata_cols[c] = names.get(c, "")
    else:
        for c in cols:
            if c in hier:
                continue
            nm = names.get(c, "")
            if nm and compact(nm) in _METADATA_HEADER_TERMS:
                region.metadata_cols[c] = nm

    if region.matrix_cols:
        for c, nm in list(region.matrix_cols.items()):
            if not nm:
                region.matrix_cols[c] = names.get(c, "")
    else:
        for c in cols:
            if c in hier or c in region.metadata_cols:
                continue
            nm = names.get(c, "")
            if not nm:
                continue  # 이름 없는 컬럼(예: 좌측 순번 열)은 어디에도 배정하지 않음
            region.matrix_cols[c] = nm


# --- 엔트리 포인트 ------------------------------------------------------------------

def detect_headers(region: "Region", canvas: "SheetCanvas", config: ParserConfig) -> None:
    cols = list(range(region.min_col, region.max_col + 1))
    title_rows = _title_rows(region)

    def content_rows(start: int, excluded: Set[int]) -> List[int]:
        return [
            r
            for r in range(start, region.max_row + 1)
            if r not in excluded and canvas.row_has_content(r, region.min_col, region.max_col)
        ]

    # key-value/주석/제목 류 region 은 헤더 탐지를 건너뛴다
    if region.region_type in _HEADERLESS_TYPES or region.role != "body":
        if not region.body_rows:
            region.body_rows = content_rows(region.min_row, title_rows)
        return

    if region.header_rows:
        # override 등으로 이미 지정 — 유지
        header_rows = sorted(region.header_rows)
        pre_rows: Set[int] = set(range(region.min_row, header_rows[0]))
    else:
        header_rows, pre_rows, best = _detect_header_rows(
            region, canvas, cols, title_rows, max(1, int(config.max_header_depth))
        )
        # autofilter 재앵커 (사용자 승인 규칙): 검출 밴드가 필터 첫 행을 놓쳤고
        # 필터행이 밴드 '아래'인 경우에만 — 필터는 작성자의 명시적 헤더 선언(코퍼스 5/5 정확).
        rf = _autofilter_first_row(canvas, region)
        if rf is not None and rf not in header_rows and (not header_rows or rf > max(header_rows)):
            rg2 = copy.copy(region)          # shallow copy — min_row 만 교체(리전 컬렉션은 read-only 사용)
            rg2.min_row = rf
            h2, p2, b2 = _detect_header_rows(
                rg2, canvas, cols, title_rows, max(1, int(config.max_header_depth))
            )
            if h2:                            # 재검출 성공 시에만 교체 — 실패 시 원본 유지(안전판)
                pre_rows = set(range(region.min_row, rf)) | p2   # 필터행 위 구간(구 밴드 포함) = 메타/pre
                header_rows, best = h2, b2
        region.header_rows = header_rows
        region.features["header_detection_score"] = float(best)
        if not header_rows:
            region.warnings.append("header_not_detected")

    parts_by_col = _collect_header_parts(canvas, cols, header_rows)
    names = _flatten_column_names(region, cols, parts_by_col)
    _assign_column_roles(region, canvas, cols, names, parts_by_col, header_rows)

    if not region.body_rows:
        if header_rows:
            excluded = title_rows | set(header_rows) | set(pre_rows)
            region.body_rows = content_rows(max(header_rows) + 1, excluded)
        else:
            region.body_rows = content_rows(region.min_row, title_rows)
