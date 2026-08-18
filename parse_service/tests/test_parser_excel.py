"""parsers/excel — 자체청킹 결과를 facade 청크 스키마로, chunk_needed=False."""
from parse_service.parsers import excel as excel_parser


def test_chunks_normalized_and_flag_false(monkeypatch):
    rag = [{"content_text": "표1 내용", "title": "시트1", "path": ["시트1"]},
           {"content_text": "표2 내용", "title": "시트2", "path": ["시트2"]}]
    monkeypatch.setattr(excel_parser, "_fetch_rag_chunks", lambda fb, fn, excel_url: (rag, {"sheets": []}))
    res = excel_parser.parse(b"PK", "a.xlsx", excel_url="http://x")
    assert res.kind == "chunks" and res.chunk_needed is False
    assert res.chunks[0]["chunk_index"] == 0
    assert res.chunks[0]["text"] == "표1 내용"
    assert "titles_context" in res.chunks[0] and "pages" in res.chunks[0]


def test_inprocess_openpyxl_smoke(monkeypatch):
    """실제 openpyxl 백엔드로 초소형 xlsx 를 파싱(파일시스템/kordoc 무관 백엔드)."""
    import io
    import openpyxl
    monkeypatch.setenv("EXCEL_PARSER_BACKEND", "openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["A2"], ws["B2"] = "이름", "값", "가", 1
    buf = io.BytesIO()
    wb.save(buf)
    res = excel_parser.parse(buf.getvalue(), "t.xlsx")
    assert res.chunk_needed is False and res.chunks
    assert res.gate_summary["parser_backend"] == "openpyxl"


def test_inprocess_uses_original_filename_as_document_title(monkeypatch):
    """:8600과 동일하게 임시파일 stem 대신 원본 basename을 문서 제목으로 전달한다."""
    from parse_service.parsers.excel.excel_parser_rag import backends, gate

    captured = {}

    class CapturingBackend:
        def parse(self, input_path, config):
            captured["temporary_stem"] = input_path.stem
            captured["document_title"] = config.document_title
            return (
                [{
                    "content_text": f"{config.document_title}의 관련 주석: 원문 주석",
                    "title": "관련 주석",
                    "path": ["관련 주석"],
                }],
                {},
            )

    monkeypatch.setattr(backends, "get_backend", lambda _name: CapturingBackend())
    monkeypatch.setattr(gate, "compute_gate_summary", lambda _path, _chunks: {"sheets": []})

    res = excel_parser.parse(
        b"PK",
        "../../2-1. 위임전결기준표(2026.04.17. 개정).xlsx",
    )

    assert captured["temporary_stem"].startswith("excel_parser_")
    assert captured["document_title"] == "2-1. 위임전결기준표(2026.04.17. 개정)"
    assert res.chunks[0]["text"].startswith(
        "2-1. 위임전결기준표(2026.04.17. 개정)의"
    )
    assert "excel_parser_" not in res.chunks[0]["text"]


def test_empty_chunks_returns_gate_summary(monkeypatch):
    """빈 청크는 더 이상 raise 하지 않는다 — gate_summary 를 실은 RouteResult 를 반환
    (깨진 엑셀은 크래시가 아니라 다운스트림 게이트에서 reject)."""
    monkeypatch.setattr(excel_parser, "_fetch_rag_chunks", lambda fb, fn, excel_url: ([], {"sheets": []}))
    res = excel_parser.parse(b"PK", "a.xlsx", excel_url="http://x")
    assert res.kind == "chunks" and res.chunks == [] and res.gate_summary is not None


def test_csv_routes_to_excel_lane():
    from parse_service import router
    from parse_service.tools import fileconvert

    assert router.domain_of("a.csv") == "excel"
    # 두 집합이 함께 바뀌어야 한다 — 하나만 바꾸면 pdf 도메인으로 떨어져
    # app.py 의 %PDF 가드가 모든 csv 를 거부한다.
    assert "csv" not in fileconvert.TEXT_EXTS
    assert fileconvert.needs_convert("a.csv") is False


def test_csv_yields_header_keyed_record_chunks(monkeypatch):
    """csv 청크는 `사번: 1001` 이어야 한다. `A: 1001` 이면 헤더 감지가 실패한 것."""
    # auto 는 전결 키워드/계층 지배도가 없으면 kordoc 으로 떨어진다 → csv 는 openpyxl 고정.
    # env 를 auto 로 두고도 성공해야 그 고정이 실제로 걸린 것이다.
    monkeypatch.setenv("EXCEL_PARSER_BACKEND", "auto")
    monkeypatch.delenv("KORDOC_BIN", raising=False)

    raw = "사번,성명,부서\n1001,김철수,전략기획부\n1002,이영희,리스크관리부\n".encode("utf-8")
    rr = excel_parser.parse(raw, "인사현황.csv")

    assert rr.kind == "chunks" and rr.chunk_needed is False
    assert rr.gate_summary is not None and rr.gate_summary.get("ok") is True
    joined = "\n".join(c["text"] for c in rr.chunks)
    assert "사번: 1001, 성명: 김철수, 부서: 전략기획부" in joined
    assert "A: 1001" not in joined


def test_csv_chunks_do_not_leak_tempfile_stem():
    raw = "사번,성명\n1001,김철수\n".encode("utf-8")
    rr = excel_parser.parse(raw, "인사현황.csv")
    joined = "\n".join(c["text"] for c in rr.chunks)
    assert "excel_parser_" not in joined
    assert "인사현황" in joined


def test_csv_formula_cell_survives_into_chunks():
    """= 로 시작하는 셀이 수식으로 승격되면 data_only 읽기에서 None 이 되어
    청크에서 조용히 사라진다(실측). 레인 왕복까지 보존되는지 본다."""
    raw = "사번,수식\n1001,=1+1\n".encode("utf-8")
    rr = excel_parser.parse(raw, "인사현황.csv")
    assert "=1+1" in "\n".join(c["text"] for c in rr.chunks)


def test_xlsx_still_honours_backend_env(monkeypatch):
    """csv 고정이 xlsx 경로의 EXCEL_PARSER_BACKEND 존중을 깨뜨리지 않았는지."""
    seen = {}

    class _FakeBackend:
        def parse(self, path, config):
            seen["backend"] = config.backend
            return [], {}

    monkeypatch.setenv("EXCEL_PARSER_BACKEND", "kordoc")
    monkeypatch.setattr(
        "parse_service.parsers.excel.excel_parser_rag.backends.get_backend",
        lambda name: _FakeBackend())
    monkeypatch.setattr(
        "parse_service.parsers.excel.excel_parser_rag.gate.compute_gate_summary",
        lambda p, c: {"ok": True, "sheets": []})

    import io as _io
    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(["a", "b"])
    buf = _io.BytesIO(); wb.save(buf)

    excel_parser.parse(buf.getvalue(), "a.xlsx")
    assert seen["backend"] == "kordoc"
