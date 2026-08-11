import io

import pytest
from openpyxl import load_workbook

from parse_service.parsers import ParserError
from parse_service.parsers.excel.csv_to_xlsx import csv_bytes_to_xlsx


def _load(raw: bytes, name: str = "인사현황.csv"):
    return load_workbook(io.BytesIO(csv_bytes_to_xlsx(raw, name)))


def test_header_row_is_styled():
    """헤더 서식이 없으면 excel_parser_rag 의 header 감지가 실패해
    `A: 1001, B: 김철수` 로 퇴화한다(detection/header_detector.py:318-323 style gate)."""
    ws = _load(b"a,b\n1,2\n").active
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb.endswith("DDDDDD")
    assert ws["A2"].font.bold in (False, None)


def test_sheet_title_is_source_stem():
    """청크 텍스트에 시트명이 박힌다 — 임시파일 stem 이 새면 검색어가 오염된다."""
    ws = _load(b"a,b\n1,2\n", "인사현황.csv").active
    assert ws.title == "인사현황"


def test_sheet_title_truncated_to_31_chars():
    """openpyxl 은 31자 초과 시트명을 거부한다."""
    ws = _load(b"a\n1\n", "가" * 40 + ".csv").active
    assert ws.title == "가" * 31


def test_sheet_title_illegal_chars_sanitized():
    """`현황[최종].csv` 같은 국내 실무 파일명이 지금은 text 레인으로 문제없이 통과한다.
    openpyxl 은 [ ] : * ? / \\ 를 시트명에서 거부하므로(ValueError → ParserError →
    문서 전체 parse_failed) 정규화하지 않으면 순수 회귀다."""
    ws = _load(b"a\n1\n", "현황[최종]:2026*3분기?.csv").active
    assert ws.title == "현황_최종__2026_3분기_"


def test_sheet_title_uses_basename_only():
    """POSIX 에서 `/` 는 진짜 경로 구분자다 — Path.stem 이 디렉터리를 떼는 게 맞고,
    실제 레인도 `safe_filename` 으로 basename 을 먼저 취한다. 백슬래시가 든
    윈도우식 이름만 정규화 대상으로 남는다."""
    assert _load(b"a\n1\n", "/tmp/sub/현황.csv").active.title == "현황"
    assert _load(b"a\n1\n", "C:\\dir\\현황.csv").active.title.endswith("현황")


def test_illegal_control_chars_stripped():
    """openpyxl 은 제어문자가 든 셀을 append 시점에 IllegalCharacterError 로 거부한다."""
    ws = _load(b"a,b\n1,bad\x07here\n").active
    assert ws["B2"].value == "badhere"


def test_cp949_and_quoted_cells():
    raw = '사번,성명\n1001,"이영희, 차장"\n'.encode("cp949")
    ws = _load(raw).active
    assert ws["A1"].value == "사번"
    assert ws["B2"].value == "이영희, 차장"


def test_quoted_newline_cell_preserved():
    """csv.reader 에 넘기는 StringIO 는 newline='' 이어야 인용 셀 안의 개행이 보존된다."""
    ws = _load('a,b\n1,"가\n나"\n'.encode("utf-8")).active
    assert ws["B2"].value == "가\n나"


def test_leading_equals_cell_is_not_a_formula():
    """`=` 로 시작하는 셀을 그대로 넣으면 openpyxl 이 수식 셀(data_type='f')로 만든다.
    엑셀 레인은 `data_only=True` 로 읽으므로 캐시된 계산값이 없는 그 셀은 **None** 이
    되어 청크에서 통째로 사라지고 게이트는 ok=True 로 통과시킨다 — 조용한 데이터 손실이
    적재까지 간다(실측). CSV 인젝션 방어도 겸한다."""
    ws = _load(b"a,b\n1,=1+1\n").active
    assert ws["B2"].data_type == "s"
    assert ws["B2"].value == "=1+1"


def test_dotfile_name_yields_usable_sheet_title():
    """`.csv` 는 Path.stem 이 '.csv' 다(확장자 없는 은닉파일 취급). 빈 시트명이 되어
    'Sheet1' 로 떨어지지 않는다는 것을 고정한다."""
    assert _load(b"a\n1\n", ".csv").active.title == ".csv"


def test_pipe_in_cell_survives():
    ws = _load("a,b\n1,리스크|관리부\n".encode("utf-8")).active
    assert ws["B2"].value == "리스크|관리부"


def test_ragged_rows_are_kept():
    """열 수가 행마다 달라도 죽지 않는다(엑셀 파서가 헤더 기준으로 맞춘다)."""
    ws = _load(b"a,b,c\n1,2\n3,4,5,6\n").active
    assert ws.max_row == 3


def test_blank_lines_skipped():
    ws = _load(b"a,b\n\n1,2\n\n").active
    assert ws.max_row == 2


def test_empty_csv_raises():
    with pytest.raises(ParserError):
        csv_bytes_to_xlsx(b"\n\n", "a.csv")
