#!/usr/bin/env python3
"""`.xls`(BIFF) 테스트 픽스처 3종을 만든다 — **재생성·검토 가능하게** 커밋한다.

    python parse_service/tests/fixtures/make_xls_fixtures.py

openpyxl 로 `.xlsx` 원본을 만들고 soffice 로 `MS Excel 97` 필터를 태워 진짜 BIFF 를 얻는다
(openpyxl 은 `.xls` 를 쓰지 못한다 — 그게 이 작업의 출발점이다).

★ `broken_formula.xls` 는 **간접 참조** 모양이어야 한다:

    J9 = `=#REF!`      ← 오류의 출처(수식 텍스트에 토큰 있음)
    H3 = `=J9`         ← 수식엔 토큰이 없고 **캐시값만 `#REF!`**

게이트는 캐시값 워크북과 수식 워크북을 하나의 ref_set 으로 합치므로
(`excel_gate.py` 의 ref_error 스캔), 통상적인 `=SUM(#REF!)` 픽스처는 LibreOffice 가
재계산으로 캐시값을 지우든 말든 **수식 스캔에서 무조건 잡힌다** → "원본 #REF! 가
보존되는가" 라는 검증이 실패할 수 없게 된다(= 아무것도 증명하지 못한다).
`H3` 는 캐시 경로로만 잡히므로 재계산으로 지워지면 실제로 빠진다.

그래서 생성 직후 아래 두 조건을 **모두** 확인하고, 하나라도 깨지면 생성을 실패시킨다:
  (a) H3 의 캐시값이 `#REF!` 다
  (b) H3 의 수식 텍스트에 `#REF!` 가 **없다**   ← (b) 가 없으면 위 함정이 되살아난다
`.xls` 는 openpyxl 로 못 읽으므로, 검사는 생성물을 soffice 로 `.xlsx` 로 되돌려서 한다.
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

HERE = Path(__file__).resolve().parent
XLS_FILTER = "xls:MS Excel 97"


def _soffice() -> str:
    for cand in ("soffice", "libreoffice", "/usr/local/bin/soffice",
                 "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        found = shutil.which(cand) if "/" not in cand else (cand if Path(cand).is_file() else None)
        if found:
            return found
    sys.exit("soffice(LibreOffice)가 필요하다 — 픽스처는 진짜 BIFF 여야 한다.")


def _convert(src: Path, out_dir: Path, target: str) -> Path:
    subprocess.run([_soffice(), "--headless", "--norestore", "--convert-to", target,
                    "--outdir", str(out_dir), str(src)],
                   check=True, capture_output=True, text=True, timeout=180)
    ext = target.split(":")[0]
    out = out_dir / f"{src.stem}.{ext}"
    if not out.is_file():
        sys.exit(f"변환 실패: {src} → {target} (산출물 없음: {list(out_dir.iterdir())})")
    return out


def _legacy_sample(ws) -> None:
    """값 + 병합셀 — 충실도 검증용."""
    ws.title = "시트1"
    ws.append(["항목", "금액", "비고"])
    ws.append(["임대료", 1200000, "월납"])
    ws.append(["관리비", 150000, ""])
    ws.append([])
    ws["A5"] = "합계 1,350,000"
    ws.merge_cells("A5:C5")


def _delegation_sample(ws) -> None:
    """"전결" 키워드 — 변환 후 openpyxl(delegation) 레인으로 라우팅되는지 확인용."""
    ws.title = "위임전결"
    ws.append(["구분", "업무내용", "대표이사", "본부장", "팀장"])
    ws.append(["일반", "1천만원 이하 지출 전결", "", "", "○"])
    ws.append(["일반", "1억원 이하 지출 전결", "", "○", ""])
    ws.append(["중요", "1억원 초과 지출 전결", "○", "", ""])


def _broken_formula(wb) -> None:
    """캐시값에만 #REF! 가 남는 간접 참조 + 정상 수식(오탐 축은 **다른 시트**에)."""
    ws = wb.active
    ws.title = "참조오류"
    ws["A1"] = "간접 참조 검사"
    ws["H3"] = "=J9"        # ← 수식엔 토큰 없음. 캐시값만 #REF!
    ws["J9"] = "=#REF!"     # ← 오류의 출처

    ok = wb.create_sheet("정상")   # 오탐 축: ref_error 가 새로 생기면 안 된다
    ok["A1"] = 10
    ok["A2"] = 20
    ok["A3"] = "=SUM(A1:A2)"


def main() -> None:
    tmp = Path(tempfile.mkdtemp(prefix="mkxls_"))
    try:
        specs = [("legacy_sample", _legacy_sample), ("delegation_sample", _delegation_sample)]
        for name, fill in specs:
            wb = Workbook()
            fill(wb.active)
            xlsx = tmp / f"{name}.xlsx"
            wb.save(xlsx)
            shutil.copy(_convert(xlsx, tmp, XLS_FILTER), HERE / f"{name}.xls")
            print(f"  ✓ {name}.xls")

        wb = Workbook()
        _broken_formula(wb)
        xlsx = tmp / "broken_formula.xlsx"
        wb.save(xlsx)
        biff = _convert(xlsx, tmp, XLS_FILTER)

        # ── 전제 검사 — .xls 는 openpyxl 로 못 읽으니 .xlsx 로 되돌려서 본다 ──
        back_dir = tmp / "back"
        back_dir.mkdir()
        back = _convert(biff, back_dir, "xlsx")
        cached = load_workbook(back, data_only=True)["참조오류"]["H3"].value
        formula = load_workbook(back, data_only=False)["참조오류"]["H3"].value
        # ⚠️ 대소문자 무시로 본다 — LibreOffice 는 캐시 오류값을 **소문자**(`#ref!`)로 쓴다
        # (2026-08-13 실측). 게이트의 ERROR_RE 도 그래서 IGNORECASE 다.
        if cached is None or "#ref!" not in str(cached).lower():
            sys.exit(f"픽스처 전제 (a) 실패 — H3 캐시값에 #REF! 가 없다: {cached!r}")
        if formula is not None and "#ref!" in str(formula).lower():
            sys.exit(
                f"픽스처 전제 (b) 실패 — H3 수식 텍스트에 #REF! 토큰이 있다: {formula!r}\n"
                "  이러면 게이트의 수식 스캔이 무조건 잡아 '재계산으로 지워지는가' 검증이 무의미해진다."
            )
        shutil.copy(biff, HERE / "broken_formula.xls")
        print(f"  ✓ broken_formula.xls  (H3 캐시={cached!r}, 수식={formula!r})")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    main()
